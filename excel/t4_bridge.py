"""
excel/t4_bridge.py - Technique 4 Excel export.

Produces a 6-sheet workbook:
  Sheet 1 "Tool Catalog" - Master list of all tools
  Sheet 2 "Domain Coverage" - Tool × Domain boolean matrix
  Sheet 3 "Feature Breakdown" - Per-tool, per-subdomain analysis
  Sheet 4 "License Distribution" - Pie chart + breakdown
  Sheet 5 "Top Multi-Domain" - Ranked platforms with coverage
  Sheet 6 "Executive Dashboard" - Strategic summary
"""

import asyncio
import json
import logging
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.domains import CYBERSECURITY_DOMAINS
from config.settings import settings
from db.t4_store import (
    get_t4_tools_with_coverage,
    get_t4_stats,
    get_t4_tool_subdomain_features_list,
)

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# Color Palette
# ═══════════════════════════════════════════════════════════════

_DARK_BG = "0E1117"
_HEADER_BG = "1E2533"
_ROW_ALT1 = "131720"
_ROW_ALT2 = "1A1F2E"
_TEXT_MAIN = "F8FAFC"
_TEXT_DIM = "94A3B8"
_ACCENT = "3B82F6"

_CHECK = "✔"
_DASH = "–"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color: str = _TEXT_MAIN, bold: bool = False, size: int = 10) -> Font:
    return Font(color=hex_color, bold=bold, size=size)


def _align(horizontal: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def _thin_border() -> Border:
    side = Side(style="thin", color="2D3748")
    return Border(left=side, right=side, top=side, bottom=side)


def _style(cell, fill_hex: str | None = None, font: Font | None = None,
           alignment: Alignment | None = None, border: bool = True) -> None:
    if fill_hex:
        cell.fill = _fill(fill_hex)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = _thin_border()


# ═══════════════════════════════════════════════════════════════
# Sheet 1: Tool Catalog
# ═══════════════════════════════════════════════════════════════

def _write_catalog_sheet(wb: Workbook, tools: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Tool Catalog")
    
    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = "Technique 4 - Tool-Level Cross-Domain Analysis"
    _style(title, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=14),
           alignment=_align("center"))
    ws.row_dimensions[1].height = 32
    
    headers = ["#", "Vendor", "Product Name", "License", "Type",
               "Domains", "Subdomains", "Total Features", "Supported", "%", "URL"]
    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        _style(cell, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True),
               alignment=_align("center"))
    ws.row_dimensions[2].height = 20
    
    for i, tool in enumerate(tools, 1):
        row_idx = i + 2
        fill_hex = _ROW_ALT1 if i % 2 == 1 else _ROW_ALT2
        
        support_pct = round(tool.get("support_rate", 0) * 100, 1)
        
        values = [
            i,
            tool.get("vendor", ""),
            tool.get("product_name", ""),
            tool.get("license_model", "Unknown"),
            (tool.get("tool_type") or "unknown").capitalize(),
            tool.get("domain_count", 0),
            tool.get("subdomain_count", 0),
            tool.get("total_subfeatures", 0),
            tool.get("supported_subfeatures", 0),
            f"{support_pct}%",
            tool.get("url", ""),
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            align = _align("center") if col_idx in (1, 4, 5, 6, 7, 8, 9, 10) else _align("left")
            
            if col_idx == 10:
                if support_pct >= 70:
                    _style(cell, fill_hex="15803D", font=_font(_TEXT_MAIN, bold=True), alignment=align)
                elif support_pct >= 40:
                    _style(cell, fill_hex="CA8A04", font=_font(_TEXT_MAIN, bold=True), alignment=align)
                else:
                    _style(cell, fill_hex=fill_hex, font=_font(_TEXT_MAIN), alignment=align)
            elif col_idx == 6 and tool.get("domain_count", 0) >= 2:
                _style(cell, fill_hex=fill_hex, font=_font(_ACCENT, bold=True), alignment=align)
            else:
                _style(cell, fill_hex=fill_hex, font=_font(_TEXT_MAIN), alignment=align)
        
        ws.row_dimensions[row_idx].height = 18
    
    widths = [5, 22, 32, 16, 12, 10, 12, 14, 12, 10, 35]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:K{len(tools) + 2}"


# ═══════════════════════════════════════════════════════════════
# Sheet 2: Domain Coverage Matrix
# ═══════════════════════════════════════════════════════════════

def _write_domain_coverage_sheet(wb: Workbook, tools: list[dict[str, Any]], all_domains: list[str]) -> None:
    ws = wb.create_sheet("Domain Coverage")
    n_domains = len(all_domains)
    total_cols = 2 + n_domains
    
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    title = ws["A1"]
    title.value = "Tool × Domain Coverage Matrix"
    _style(title, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=12),
           alignment=_align("center"))
    ws.row_dimensions[1].height = 26
    
    ws.cell(row=2, column=1, value="Vendor")
    ws.cell(row=2, column=2, value="Product Name")
    _style(ws.cell(row=2, column=1), fill_hex=_HEADER_BG,
           font=_font(_TEXT_MAIN, bold=True), alignment=_align("center"))
    _style(ws.cell(row=2, column=2), fill_hex=_HEADER_BG,
           font=_font(_TEXT_MAIN, bold=True), alignment=_align("left"))
    
    for di, domain in enumerate(all_domains, 3):
        cell = ws.cell(row=2, column=di, value=domain[:15])
        _style(cell, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=8),
               alignment=_align("center", wrap=True))
        ws.column_dimensions[get_column_letter(di)].width = 12
    ws.row_dimensions[2].height = 38
    
    for i, tool in enumerate(tools, 1):
        row_idx = i + 2
        fill_hex = _ROW_ALT1 if i % 2 == 1 else _ROW_ALT2
        
        tool_domains = set(tool.get("domain_list", []))
        
        c1 = ws.cell(row=row_idx, column=1, value=tool.get("vendor", ""))
        _style(c1, fill_hex=fill_hex, font=_font(_TEXT_DIM, size=9), alignment=_align("left"))
        c2 = ws.cell(row=row_idx, column=2, value=tool.get("product_name", ""))
        _style(c2, fill_hex=fill_hex, font=_font(_TEXT_MAIN), alignment=_align("left"))
        
        for di, domain in enumerate(all_domains, 3):
            has_domain = domain in tool_domains
            bg_color = _ACCENT if has_domain else fill_hex
            cell = ws.cell(row=row_idx, column=di, value=_CHECK if has_domain else _DASH)
            _style(cell, fill_hex=bg_color, font=_font(_TEXT_MAIN, bold=has_domain),
                   alignment=_align("center"))
        
        ws.row_dimensions[row_idx].height = 16
    
    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32


# ═══════════════════════════════════════════════════════════════
# Sheet 3: Feature Breakdown
# ═══════════════════════════════════════════════════════════════

def _write_feature_breakdown_sheet(wb: Workbook, tools: list[dict[str, Any]],
                                    tool_subdomain_data: dict[int, list[dict]]) -> None:
    ws = wb.create_sheet("Feature Breakdown")
    
    ws.merge_cells("A1:J1")
    title = ws["A1"]
    title.value = "Per-Tool, Per-Subdomain Feature Support"
    _style(title, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=12),
           alignment=_align("center"))
    ws.row_dimensions[1].height = 26
    
    headers = ["Vendor", "Product Name", "Subdomain", "Domain",
               "Total", "Supported", "Partial", "%", "Level"]
    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        _style(cell, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True),
               alignment=_align("center"))
    ws.row_dimensions[2].height = 20
    
    row_idx = 3
    for tool in tools:
        tool_id = tool.get("id")
        breakdown = tool_subdomain_data.get(tool_id, [])
        
        if not breakdown:
            continue
        
        for i, entry in enumerate(breakdown):
            fill_hex = _ROW_ALT1 if row_idx % 2 == 1 else _ROW_ALT2
            
            values = [
                tool.get("vendor", ""),
                tool.get("product_name", ""),
                entry.get("subdomain_name", ""),
                entry.get("domain_name", ""),
                entry.get("total_subfeatures", 0),
                entry.get("supported_subfeatures", 0),
                entry.get("partial_subfeatures", 0),
                f"{float(entry.get('support_pct', 0) or 0):.1f}%",
                entry.get("support_level", "Unknown"),
            ]
            
            for col_idx_val, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx_val, value=value)
                align = _align("center") if col_idx_val in (5, 6, 7, 8) else _align("left")
                
                if col_idx_val == 9:
                    level = entry.get("support_level", "Unknown")
                    if level == "High":
                        _style(cell, fill_hex="15803D", font=_font(_TEXT_MAIN, bold=True), alignment=align)
                    elif level == "Medium":
                        _style(cell, fill_hex="CA8A04", font=_font(_TEXT_MAIN), alignment=align)
                    else:
                        _style(cell, fill_hex="B91C1C", font=_font(_TEXT_MAIN), alignment=align)
                else:
                    _style(cell, fill_hex=fill_hex, font=_font(_TEXT_MAIN), alignment=align)
            
            row_idx += 1
    
    widths = [22, 32, 28, 24, 10, 12, 12, 10, 12]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.freeze_panes = "C3"


# ═══════════════════════════════════════════════════════════════
# Sheet 4: License Distribution
# ═══════════════════════════════════════════════════════════════

def _write_license_sheet(wb: Workbook, stats: dict[str, Any]) -> None:
    ws = wb.create_sheet("License Distribution")
    
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "License Model Distribution"
    _style(title, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=13),
           alignment=_align("center"))
    ws.row_dimensions[1].height = 28
    
    headers = ["License Model", "Count", "Percentage", "Notes"]
    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        _style(cell, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True),
               alignment=_align("center"))
    ws.row_dimensions[2].height = 18
    
    license_counts = stats.get("license_counts", {})
    total = sum(license_counts.values()) if license_counts else 1
    
    row_idx = 3
    for lic in ["Commercial", "GPL-3.0", "GPL-2.0", "LGPL", "MIT", "BSD",
                "Apache-2.0", "MPL-2.0", "Freemium", "Proprietary", "Unknown"]:
        count = license_counts.get(lic, 0)
        if count == 0:
            continue
        
        pct = (count / total * 100) if total > 0 else 0
        fill_hex = _ROW_ALT1 if row_idx % 2 == 1 else _ROW_ALT2
        
        note = ""
        if lic == "Commercial":
            note = "Proprietary, enterprise licenses"
        elif "GPL" in lic:
            note = "Copyleft open source"
        elif lic in ("MIT", "BSD", "Apache-2.0"):
            note = "Permissive open source"
        elif lic == "Freemium":
            note = "Free tier + paid premium"
        
        values = [lic, count, f"{pct:.1f}%", note]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            align = _align("center") if col_idx in (2, 3) else _align("left")
            _style(cell, fill_hex=fill_hex, font=_font(_TEXT_MAIN), alignment=align)
        
        row_idx += 1
    
    widths = [18, 10, 14, 35]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    data_ws = wb.create_sheet("_T4_LicenseChartData")
    data_ws.sheet_state = "hidden"
    data_ws["A1"] = "License"
    data_ws["B1"] = "Count"
    row = 2
    for lic, count in license_counts.items():
        if count > 0:
            data_ws[f"A{row}"] = lic
            data_ws[f"B{row}"] = count
            row += 1
    
    if row > 2:
        pie = PieChart()
        pie.title = "License Distribution"
        labels = Reference(data_ws, min_col=1, min_row=2, max_row=row - 1)
        data = Reference(data_ws, min_col=2, min_row=1, max_row=row - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 18
        pie.height = 12
        ws.add_chart(pie, "F2")


# ═══════════════════════════════════════════════════════════════
# Sheet 5: Top Multi-Domain Platforms
# ═══════════════════════════════════════════════════════════════

def _write_top_platforms_sheet(wb: Workbook, tools: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Top Multi-Domain")
    
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "Top Multi-Domain Platform Tools"
    _style(title, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=13),
           alignment=_align("center"))
    ws.row_dimensions[1].height = 28
    
    headers = ["Rank", "Vendor", "Product Name", "License", "Domains", "Subdomains",
               "Support %", "Domain Coverage"]
    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        _style(cell, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True),
               alignment=_align("center"))
    ws.row_dimensions[2].height = 18
    
    multi_domain_tools = [t for t in tools if t.get("domain_count", 0) >= 2]
    multi_domain_tools.sort(key=lambda x: x.get("domain_count", 0), reverse=True)
    
    max_platforms = getattr(settings, "t4_excel_max_platforms", 50)
    for i, tool in enumerate(multi_domain_tools[:max_platforms], 1):
        row_idx = i + 2
        fill_hex = _ROW_ALT1 if i % 2 == 1 else _ROW_ALT2
        
        domain_list = tool.get("domain_list", [])
        domain_str = ", ".join(domain_list[:3])
        if len(domain_list) > 3:
            domain_str += f" ... +{len(domain_list) - 3}"
        
        support_pct = round(tool.get("support_rate", 0) * 100, 1)
        
        values = [
            i,
            tool.get("vendor", ""),
            tool.get("product_name", ""),
            tool.get("license_model", "Unknown"),
            tool.get("domain_count", 0),
            tool.get("subdomain_count", 0),
            f"{support_pct}%",
            domain_str,
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            align = _align("center") if col_idx in (1, 4, 5, 6, 7) else _align("left")
            _style(cell, fill_hex=fill_hex, font=_font(_TEXT_MAIN), alignment=align)
        
        ws.row_dimensions[row_idx].height = 18
    
    widths = [8, 22, 32, 16, 12, 14, 12, 50]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════
# Sheet 6: Executive Dashboard
# ═══════════════════════════════════════════════════════════════

def _write_executive_dashboard_sheet(wb: Workbook, stats: dict[str, Any],
                                       tools: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Executive Dashboard")
    ws.sheet_properties.tabColor = "3B82F6"
    
    ws.sheet_view.showGridLines = False
    
    for r in range(1, 35):
        for c in range(1, 18):
            _style(ws.cell(row=r, column=c), fill_hex=_DARK_BG, border=False)
    
    ws.merge_cells("B2:Q3")
    title = ws["B2"]
    title.value = "Technique 4: Tool-Level Cross-Domain Analysis"
    _style(title, fill_hex=_DARK_BG, font=_font(_TEXT_MAIN, bold=True, size=20),
           alignment=_align("left"))
    
    ws.merge_cells("B5:H10")
    summary = ws["B5"]
    summary.value = (
        f"Total Tools: {stats.get('total', 0)}\n"
        f"Enterprise: {stats.get('enterprise', 0)}\n"
        f"Open Source: {stats.get('opensource', 0)}\n"
        f"Multi-Domain Tools: {stats.get('multi_domain', 0)}\n"
        f"Average Support Rate: {stats.get('avg_support_rate', 0) * 100:.1f}%"
    )
    _style(summary, fill_hex=_ROW_ALT1, font=_font(_TEXT_DIM, size=11),
           alignment=_align("left", wrap=True), border=True)
    
    data_ws = wb.create_sheet("_T4_DashboardData")
    data_ws.sheet_state = "hidden"
    
    vendor_counts = Counter([t.get("vendor", "Unknown") for t in tools])
    top_vendors = vendor_counts.most_common(10)
    
    data_ws["A1"] = "Vendor"
    data_ws["B1"] = "Count"
    row = 2
    for vendor, count in top_vendors:
        data_ws[f"A{row}"] = vendor
        data_ws[f"B{row}"] = count
        row += 1
    
    if row > 2:
        bar = BarChart()
        bar.type = "col"
        bar.title = "Top 10 Vendors by Tool Count"
        bar.style = 10
        bar.width = 18
        bar.height = 12
        labels = Reference(data_ws, min_col=1, min_row=2, max_row=row - 1)
        data = Reference(data_ws, min_col=2, min_row=1, max_row=row - 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(labels)
        bar.x_axis.tickLblPos = "low"
        ws.add_chart(bar, "B12")
    
    license_counts = stats.get("license_counts", {})
    data_ws["D1"] = "License"
    data_ws["E1"] = "Count"
    row = 2
    for lic, count in license_counts.items():
        if count > 0:
            data_ws[f"D{row}"] = lic
            data_ws[f"E{row}"] = count
            row += 1
    
    if row > 2:
        pie = PieChart()
        pie.title = "License Distribution"
        pie.style = 10
        pie.width = 16
        pie.height = 12
        labels = Reference(data_ws, min_col=4, min_row=2, max_row=row - 1)
        data = Reference(data_ws, min_col=5, min_row=1, max_row=row - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, "L12")


# ═══════════════════════════════════════════════════════════════
# Main Export Function
# ═══════════════════════════════════════════════════════════════

def _sync_write_t4_excel(
    path: Path,
    tools: list[dict[str, Any]],
    stats: dict[str, Any],
    tool_subdomain_data: dict[int, list[dict]],
    all_domains: list[str],
) -> None:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    _write_executive_dashboard_sheet(wb, stats, tools)
    _write_catalog_sheet(wb, tools)
    _write_domain_coverage_sheet(wb, tools, all_domains)
    _write_feature_breakdown_sheet(wb, tools, tool_subdomain_data)
    _write_license_sheet(wb, stats)
    _write_top_platforms_sheet(wb, tools)
    
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    logger.info(f"T4 Excel workbook saved → {path} ({len(tools)} tools, 6 sheets)")


async def export_t4_workbook() -> str:
    """
    Export the full T4 analysis workbook.
    Fetches data, writes 6-sheet Excel, returns output path.
    """
    from db.t4_store import get_all_t4_tool_subdomain_features
    
    tools = await get_t4_tools_with_coverage(min_domain_count=1)
    
    if not tools:
        raise ValueError("No T4 tools found. Run the T4 analysis pipeline first.")
    
    stats = await get_t4_stats()
    
    all_domains = list(CYBERSECURITY_DOMAINS)
    
    tool_subdomain_data = await get_all_t4_tool_subdomain_features()
    
    output_path = Path(settings.t4_excel_output_path)
    
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(
        None,
        _sync_write_t4_excel,
        output_path,
        tools,
        stats,
        tool_subdomain_data,
        all_domains,
    )
    
    return str(output_path)
