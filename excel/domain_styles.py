"""
Technique 2: Excel styling for domain-level tool rankings.
Separate from Technique 1 styles.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPPORT_FULL = "✔"
SUPPORT_NONE = "✘"
SUPPORT_PARTIAL = "Partial"

T2_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
T2_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
T2_FILL_AMBER = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

T2_FILL_TITLE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
T2_FILL_ENTERPRISE = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
T2_FILL_OPENSOURCE = PatternFill(start_color="538135", end_color="538135", fill_type="solid")
T2_FILL_HEADER = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
T2_FILL_SCORE_HIGH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
T2_FILL_SCORE_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
T2_FILL_SCORE_LOW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

T2_FEATURE_PALETTE: list[PatternFill] = [
    PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid"),
    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
    PatternFill(start_color="E8F4EA", end_color="E8F4EA", fill_type="solid"),
]

T2_FONT_TITLE = Font(bold=True, color="FFFFFF", size=14)
T2_FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
T2_FONT_SUBHEADER = Font(bold=True, color="000000", size=10)
T2_FONT_TOOL = Font(bold=False, color="000000", size=10)
T2_FONT_RANK = Font(bold=True, color="000000", size=10)
T2_FONT_SCORE_HIGH = Font(bold=True, color="006100", size=10)
T2_FONT_SCORE_MED = Font(bold=True, color="9C5700", size=10)
T2_FONT_SCORE_LOW = Font(bold=True, color="9C0006", size=10)
T2_FONT_SUPPORT = Font(bold=True, color="000000", size=10)
T2_FONT_LEGEND = Font(bold=False, color="000000", size=9)

T2_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
T2_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
T2_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

_THIN = Side(style="thin")
T2_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def t2_get_support_fill(value: str) -> PatternFill:
    return {
        SUPPORT_FULL: T2_FILL_GREEN,
        SUPPORT_NONE: T2_FILL_RED,
        SUPPORT_PARTIAL: T2_FILL_AMBER,
    }.get(value, T2_FILL_RED)


def t2_get_score_fill(score: float) -> PatternFill:
    if score >= 70:
        return T2_FILL_SCORE_HIGH
    elif score >= 50:
        return T2_FILL_SCORE_MED
    else:
        return T2_FILL_SCORE_LOW


def t2_get_score_font(score: float) -> Font:
    if score >= 70:
        return T2_FONT_SCORE_HIGH
    elif score >= 50:
        return T2_FONT_SCORE_MED
    else:
        return T2_FONT_SCORE_LOW


def t2_style_cell(cell, *, font=None, fill=None, alignment=None, border=T2_BORDER_THIN) -> None:
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def t2_set_column_widths(ws, num_tool_cols: int) -> None:
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    for i in range(4, 4 + num_tool_cols):
        ws.column_dimensions[get_column_letter(i)].width = 14
