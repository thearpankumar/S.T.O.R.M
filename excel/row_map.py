import logging
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook, load_workbook

from config.settings import settings

logger = logging.getLogger(__name__)

ROW_MAP_SHEET = "__row_map__"


@dataclass
class RowMapEntry:
    subdomain: str
    sheet_name: str
    data_start_row: int
    data_end_row: int
    feature: str | None = None


async def get_row_map_entries(workbook_path: str | None = None) -> list[RowMapEntry]:
    """
    Read all row map entries from the hidden __row_map__ sheet.
    """
    from pathlib import Path
    import asyncio
    
    path = Path(workbook_path or settings.excel_output_path)
    
    if not path.exists():
        return []
    
    def _sync_read() -> list[RowMapEntry]:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        
        if ROW_MAP_SHEET not in wb.sheetnames:
            wb.close()
            return []
        
        ws = wb[ROW_MAP_SHEET]
        entries = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            entries.append(RowMapEntry(
                subdomain=row[0] or "",
                sheet_name=row[1] or "",
                data_start_row=int(row[2]) if row[2] else 0,
                data_end_row=int(row[3]) if row[3] else 0,
                feature=row[4] if len(row) > 4 else None
            ))
        
        wb.close()
        return entries
    
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, _sync_read)


async def save_row_map_entry(
    entry: RowMapEntry,
    workbook_path: str | None = None
) -> None:
    """
    Save or update a row map entry in the hidden __row_map__ sheet.
    """
    from pathlib import Path
    import asyncio
    
    path = Path(workbook_path or settings.excel_output_path)
    
    def _sync_write() -> None:
        if path.exists():
            wb = load_workbook(str(path))
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
        
        if ROW_MAP_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(ROW_MAP_SHEET)
            ws.append(["subdomain", "sheet_name", "data_start_row", "data_end_row", "feature"])
            ws.sheet_state = "hidden"
        else:
            ws = wb[ROW_MAP_SHEET]
        
        existing_row = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == entry.subdomain and (not entry.feature or row[4] == entry.feature):
                existing_row = idx
                break
        
        if existing_row:
            ws.cell(row=existing_row, column=1, value=entry.subdomain)
            ws.cell(row=existing_row, column=2, value=entry.sheet_name)
            ws.cell(row=existing_row, column=3, value=entry.data_start_row)
            ws.cell(row=existing_row, column=4, value=entry.data_end_row)
            ws.cell(row=existing_row, column=5, value=entry.feature)
        else:
            ws.append([
                entry.subdomain,
                entry.sheet_name,
                entry.data_start_row,
                entry.data_end_row,
                entry.feature
            ])
        
        wb.save(str(path))
        wb.close()
    
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, _sync_write)


async def get_data_range_for_subdomain(
    subdomain: str,
    workbook_path: str | None = None
) -> tuple[int, int] | None:
    """
    Get the data row range for a given subdomain.
    Returns (start_row, end_row) or None if not found.
    """
    entries = await get_row_map_entries(workbook_path)
    
    for entry in entries:
        if entry.subdomain == subdomain:
            return entry.data_start_row, entry.data_end_row
    
    return None


async def update_data_range_for_subdomain(
    subdomain: str,
    data_start_row: int,
    data_end_row: int,
    workbook_path: str | None = None
) -> None:
    """
    Update the data row range for a given subdomain.
    """
    entry = RowMapEntry(
        subdomain=subdomain,
        sheet_name=f"{subdomain} Matrix",
        data_start_row=data_start_row,
        data_end_row=data_end_row
    )
    await save_row_map_entry(entry, workbook_path)
