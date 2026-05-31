import asyncio
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from config.settings import settings
from excel.styles import (
    FILL_TITLE, FILL_ENTERPRISE, FILL_OPENSOURCE,
    FILL_SUBDOMAIN, FILL_SUBFEATURE,
    FONT_TITLE, FONT_COL_HEADER,
    FONT_SUBDOMAIN, FONT_FEATURE, FONT_SUBFEATURE,
    ALIGN_CENTER, ALIGN_LEFT,
    BORDER_THIN,
    style_cell,
)
from db.subdomain_store import get_t2_subdomain_tools, get_eligible_subdomains

logger = logging.getLogger(__name__)

_t2_excel_lock: asyncio.Lock | None = None


def _get_t2_excel_lock() -> asyncio.Lock:
    global _t2_excel_lock
    if _t2_excel_lock is None:
        _t2_excel_lock = asyncio.Lock()
    return _t2_excel_lock


def _write_t2_header_rows(ws: Any, title: str) -> None:
    """Row 1: title span. Row 2: column headers."""
    # Title cell A1:E1
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = title
    style_cell(cell, font=FONT_TITLE, fill=FILL_TITLE, alignment=ALIGN_CENTER)
    
    # Style the merged-over cells for borders
    for col in range(2, 6):
        style_cell(ws.cell(row=1, column=col), fill=FILL_TITLE, border=BORDER_THIN)
        
    ws.row_dimensions[1].height = 30

    # Column headers
    headers = ["Rank", "Vendor", "Product Name", "Type", "Composite Score"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = label
        style_cell(cell, font=FONT_COL_HEADER, fill=FILL_TITLE, alignment=ALIGN_CENTER)

    ws.row_dimensions[2].height = 20


def _write_t2_data_rows(ws: Any, tools: list[dict[str, Any]]) -> None:
    """Write tool ranking data."""
    row_idx = 3
    
    # Generate continuous numbering 1 to N, sorted by score
    sorted_tools = sorted(tools, key=lambda x: x.get("composite_score", 0), reverse=True)
    
    for i, tool in enumerate(sorted_tools, start=1):
        vendor = tool.get("vendor", "")
        product_name = tool.get("product_name", "")
        tool_type = tool.get("tool_type", "").capitalize()
        score = tool.get("composite_score", 0.0)
        
        # We alternate fill slightly or just use SUBFEATURE fill (white) and SUBDOMAIN (light grey)
        # Let's use Enterprise/OpenSource specific colors for the Type column if desired,
        # but for simplicity, we'll just use a clean alternating white/light-grey row style.
        fill = FILL_SUBDOMAIN if i % 2 == 1 else FILL_SUBFEATURE
        
        # Rank
        c1 = ws.cell(row=row_idx, column=1, value=i)
        style_cell(c1, font=FONT_SUBDOMAIN, fill=fill, alignment=ALIGN_CENTER)
        
        # Vendor
        c2 = ws.cell(row=row_idx, column=2, value=vendor)
        style_cell(c2, font=FONT_SUBFEATURE, fill=fill, alignment=ALIGN_LEFT)
        
        # Product Name
        c3 = ws.cell(row=row_idx, column=3, value=product_name)
        style_cell(c3, font=FONT_FEATURE, fill=fill, alignment=ALIGN_LEFT)
        
        # Type
        c4 = ws.cell(row=row_idx, column=4, value=tool_type)
        type_font = FONT_SUBDOMAIN
        style_cell(c4, font=type_font, fill=fill, alignment=ALIGN_CENTER)
        
        # Score
        c5 = ws.cell(row=row_idx, column=5, value=f"{score:.1f}")
        style_cell(c5, font=FONT_FEATURE, fill=fill, alignment=ALIGN_CENTER)
        
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    # Set column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15


def _sync_write_t2_excel(
    path: Path,
    subdomain_name: str,
    tools: list[dict[str, Any]],
) -> None:
    if path.exists():
        wb = load_workbook(str(path))
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_name = f"{subdomain_name[:21]} Rankings"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    _write_t2_header_rows(ws, f"{subdomain_name} Tool Rankings (T2)")
    _write_t2_data_rows(ws, tools)

    wb.save(str(path))
    logger.info(f"Saved T2 Excel workbook to {path}")


async def create_t2_workbook_from_db(subdomain_id: int, subdomain_name: str) -> str:
    tools = await get_t2_subdomain_tools(subdomain_id)
    
    if not tools:
        logger.warning(f"No T2 tools found for {subdomain_name}")
        raise ValueError(f"No ranked tools available for {subdomain_name}")

    # Determine output path for T2. We append '_T2_Rankings' to the base filename.
    base_path = Path(settings.excel_output_path)
    output_path = base_path.with_name(f"{base_path.stem}_T2_Rankings{base_path.suffix}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    async with _get_t2_excel_lock():
        await asyncio.get_running_loop().run_in_executor(
            None,
            _sync_write_t2_excel,
            output_path,
            subdomain_name,
            tools,
        )
        
    return str(output_path)


def _sync_export_all_t2(
    path: Path,
    subdomains_data: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary", 0)
    summary_headers = ["Domain", "Subdomain", "Tools Ranked"]
    for col_idx, label in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.value = label
        style_cell(cell, font=FONT_COL_HEADER, fill=FILL_TITLE, alignment=ALIGN_CENTER)

    row_idx = 2
    for sd_data in subdomains_data:
        ws_summary.cell(row=row_idx, column=1).value = sd_data.get("domain", "")
        ws_summary.cell(row=row_idx, column=2).value = sd_data.get("name", "")
        ws_summary.cell(row=row_idx, column=3).value = sd_data.get("tools_count", 0)
        row_idx += 1

    for col in range(1, 4):
        ws_summary.column_dimensions[get_column_letter(col)].width = 30

    # ── Per-subdomain sheets ───────────────────────────────────────────────
    for sd_data in subdomains_data:
        subdomain_name = sd_data.get("name", "")
        tools          = sd_data.get("tools", [])

        if not tools:
            continue

        sheet_name = f"{subdomain_name[:21]} Rankings"
        ws = wb.create_sheet(sheet_name)

        title = f"{subdomain_name} Tool Rankings (T2)"

        _write_t2_header_rows(ws, title)
        _write_t2_data_rows(ws, tools)

    wb.save(str(path))
    logger.info(f"Exported all T2 subdomains to {path}")


async def export_all_t2_subdomains() -> str:
    base_path = Path(settings.excel_output_path)
    output_path = base_path.with_name(f"{base_path.stem}_T2_Rankings{base_path.suffix}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    eligible = await get_eligible_subdomains()
    subdomains_data = []

    for sd in eligible:
        tools = await get_t2_subdomain_tools(sd["id"])
        if tools:
            subdomains_data.append({
                "domain": sd.get("domain_name", ""),
                "name": sd.get("name", ""),
                "tools_count": len(tools),
                "tools": tools,
            })

    if not subdomains_data:
        raise ValueError("No ranked tools available to export across any subdomains.")

    async with _get_t2_excel_lock():
        await asyncio.get_running_loop().run_in_executor(
            None,
            _sync_export_all_t2,
            output_path,
            subdomains_data,
        )

    return str(output_path)
