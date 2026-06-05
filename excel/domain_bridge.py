"""
Technique 2: Excel Export for Domain Tool Rankings.

Generates a separate workbook (technique2_domain_rankings.xlsx) with:
- Summary sheet
- One sheet per domain with ranked tools and feature matrix
- Legend and methodology sheets
"""

import asyncio
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config.settings import settings
from excel.domain_styles import (
    T2_FILL_TITLE, T2_FILL_ENTERPRISE, T2_FILL_OPENSOURCE,
    T2_FILL_HEADER, T2_FILL_GREEN, T2_FILL_RED, T2_FILL_AMBER,
    T2_FONT_TITLE, T2_FONT_HEADER, T2_FONT_SUBHEADER, T2_FONT_TOOL,
    T2_FONT_RANK, T2_FONT_SUPPORT, T2_FONT_LEGEND,
    T2_ALIGN_CENTER, T2_ALIGN_LEFT, T2_ALIGN_RIGHT,
    T2_BORDER_THIN, T2_FEATURE_PALETTE,
    t2_get_support_fill, t2_get_score_fill, t2_get_score_font,
    t2_style_cell, t2_set_column_widths,
    SUPPORT_FULL, SUPPORT_NONE, SUPPORT_PARTIAL,
)
from db.domain_store import (
    get_all_t2_domain_rankings,
    get_t2_domain_tools,
    get_t2_domain_features,
    get_t2_domain_subfeatures_by_domain,
    get_t2_domain_matrix_cells,
    get_t2_domain_ranking,
)

logger = logging.getLogger(__name__)

_excel_lock: asyncio.Lock | None = None


def _get_excel_lock() -> asyncio.Lock:
    global _excel_lock
    if _excel_lock is None:
        _excel_lock = asyncio.Lock()
    return _excel_lock


def _write_summary_sheet(wb: Workbook, rankings: list[dict]) -> None:
    ws = wb.create_sheet("Summary", 0)
    
    headers = ["Domain", "Enterprise Tools", "OSS Tools", "Top Enterprise", "Score", "Top OSS", "Score", "Status"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = label
        t2_style_cell(cell, font=T2_FONT_HEADER, fill=T2_FILL_TITLE, alignment=T2_ALIGN_CENTER)
    
    row_idx = 2
    for ranking in rankings:
        ws.cell(row=row_idx, column=1).value = ranking.get("domain_name", "")
        ws.cell(row=row_idx, column=2).value = f"{ranking.get('total_enterprise_tools', 0)}→{ranking.get('selected_enterprise_tools', 0)}"
        ws.cell(row=row_idx, column=3).value = f"{ranking.get('total_opensource_tools', 0)}→{ranking.get('selected_opensource_tools', 0)}"
        
        top_ent = ranking.get("top_enterprise", {})
        ws.cell(row=row_idx, column=4).value = top_ent.get("name", "-")
        ws.cell(row=row_idx, column=5).value = top_ent.get("score", "-")
        
        top_oss = ranking.get("top_opensource", {})
        ws.cell(row=row_idx, column=6).value = top_oss.get("name", "-")
        ws.cell(row=row_idx, column=7).value = top_oss.get("score", "-")
        
        status = ranking.get("status", "pending")
        ws.cell(row=row_idx, column=8).value = status
        status_fill = T2_FILL_GREEN if status == "done" else T2_FILL_AMBER
        t2_style_cell(ws.cell(row=row_idx, column=8), fill=status_fill, alignment=T2_ALIGN_CENTER)
        
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = T2_BORDER_THIN
        
        row_idx += 1
    
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_domain_sheet(
    wb: Workbook,
    domain_name: str,
    tools_enterprise: list[dict],
    tools_opensource: list[dict],
    features: list[dict],
    subfeatures: list[dict],
    matrix_cells: list[dict],
) -> None:
    sheet_name = domain_name[:30]
    ws = wb.create_sheet(sheet_name)
    
    _write_ranking_section(ws, domain_name, tools_enterprise, tools_opensource)
    
    if subfeatures and matrix_cells:
        start_row = len(tools_enterprise) + len(tools_opensource) + 8
        _write_matrix_section(ws, subfeatures, tools_enterprise, tools_opensource, matrix_cells, start_row)


def _write_ranking_section(
    ws,
    domain_name: str,
    tools_enterprise: list[dict],
    tools_opensource: list[dict],
) -> None:
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = f"{domain_name} - Top Ranked Tools"
    t2_style_cell(cell, font=T2_FONT_TITLE, fill=T2_FILL_TITLE, alignment=T2_ALIGN_CENTER)
    ws.row_dimensions[1].height = 28
    
    headers = ["#", "Tool Name", "Score", "Presence", "Coverage", "Market", "Rank Dist"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = label
        t2_style_cell(cell, font=T2_FONT_SUBHEADER, fill=T2_FILL_HEADER, alignment=T2_ALIGN_CENTER)
    
    row_idx = 3
    if tools_enterprise:
        ws.merge_cells(f"A{row_idx}:G{row_idx}")
        cell = ws.cell(row=row_idx, column=1)
        cell.value = "ENTERPRISE"
        t2_style_cell(cell, font=T2_FONT_HEADER, fill=T2_FILL_ENTERPRISE, alignment=T2_ALIGN_CENTER)
        row_idx += 1
        
        for tool in tools_enterprise:
            rank = tool.get("rank_position", 0)
            score = tool.get("composite_score", 0)
            
            ws.cell(row=row_idx, column=1).value = rank
            t2_style_cell(ws.cell(row=row_idx, column=1), font=T2_FONT_RANK, alignment=T2_ALIGN_CENTER)
            
            ws.cell(row=row_idx, column=2).value = tool.get("product_name", "")
            t2_style_cell(ws.cell(row=row_idx, column=2), font=T2_FONT_TOOL, alignment=T2_ALIGN_LEFT)
            
            ws.cell(row=row_idx, column=3).value = f"{score:.1f}"
            t2_style_cell(
                ws.cell(row=row_idx, column=3),
                font=t2_get_score_font(score),
                fill=t2_get_score_fill(score),
                alignment=T2_ALIGN_CENTER
            )
            
            ws.cell(row=row_idx, column=4).value = f"{tool.get('subdomain_presence_score', 0)*100:.0f}%"
            ws.cell(row=row_idx, column=5).value = f"{tool.get('feature_coverage_score', 0)*100:.0f}%"
            ws.cell(row=row_idx, column=6).value = f"{tool.get('market_presence_score', 0)*100:.0f}%"
            ws.cell(row=row_idx, column=7).value = f"{tool.get('rank_distribution_score', 0)*100:.0f}%"
            
            for col in range(4, 8):
                t2_style_cell(ws.cell(row=row_idx, column=col), alignment=T2_ALIGN_CENTER)
            
            row_idx += 1
    
    if tools_opensource:
        ws.merge_cells(f"A{row_idx}:G{row_idx}")
        cell = ws.cell(row=row_idx, column=1)
        cell.value = "OPEN-SOURCE"
        t2_style_cell(cell, font=T2_FONT_HEADER, fill=T2_FILL_OPENSOURCE, alignment=T2_ALIGN_CENTER)
        row_idx += 1
        
        for tool in tools_opensource:
            rank = tool.get("rank_position", 0)
            score = tool.get("composite_score", 0)
            
            ws.cell(row=row_idx, column=1).value = f"OSS{rank}"
            t2_style_cell(ws.cell(row=row_idx, column=1), font=T2_FONT_RANK, alignment=T2_ALIGN_CENTER)
            
            ws.cell(row=row_idx, column=2).value = tool.get("product_name", "")
            t2_style_cell(ws.cell(row=row_idx, column=2), font=T2_FONT_TOOL, alignment=T2_ALIGN_LEFT)
            
            ws.cell(row=row_idx, column=3).value = f"{score:.1f}"
            t2_style_cell(
                ws.cell(row=row_idx, column=3),
                font=t2_get_score_font(score),
                fill=t2_get_score_fill(score),
                alignment=T2_ALIGN_CENTER
            )
            
            ws.cell(row=row_idx, column=4).value = f"{tool.get('subdomain_presence_score', 0)*100:.0f}%"
            ws.cell(row=row_idx, column=5).value = f"{tool.get('feature_coverage_score', 0)*100:.0f}%"
            ws.cell(row=row_idx, column=6).value = f"{tool.get('market_presence_score', 0)*100:.0f}%"
            ws.cell(row=row_idx, column=7).value = f"{tool.get('rank_distribution_score', 0)*100:.0f}%"
            
            for col in range(4, 8):
                t2_style_cell(ws.cell(row=row_idx, column=col), alignment=T2_ALIGN_CENTER)
            
            row_idx += 1
    
    t2_set_column_widths(ws, 0)


def _write_matrix_section(
    ws,
    subfeatures: list[dict],
    tools_enterprise: list[dict],
    tools_opensource: list[dict],
    matrix_cells: list[dict],
    start_row: int,
) -> None:
    all_tools = tools_enterprise + tools_opensource
    
    if not all_tools or not subfeatures:
        return
    
    ws.merge_cells(f"A{start_row}:G{start_row}")
    cell = ws.cell(row=start_row, column=1)
    cell.value = "Feature Matrix"
    t2_style_cell(cell, font=T2_FONT_TITLE, fill=T2_FILL_TITLE, alignment=T2_ALIGN_CENTER)
    start_row += 1
    
    headers = ["Feature", "Subfeature"] + [t["product_name"][:12] for t in all_tools]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = label
        fill = T2_FILL_HEADER
        if col_idx > 2:
            tool_idx = col_idx - 3
            if tool_idx < len(tools_enterprise):
                fill = T2_FILL_ENTERPRISE
            else:
                fill = T2_FILL_OPENSOURCE
        t2_style_cell(cell, font=T2_FONT_SUBHEADER, fill=fill, alignment=T2_ALIGN_CENTER)
    
    start_row += 1
    
    sf_by_feature: dict[str, list[dict]] = {}
    for sf in subfeatures:
        feat = sf.get("feature_name", "")
        sf_by_feature.setdefault(feat, []).append(sf)
    
    feature_color_idx = 0
    for feature_name, feature_subfeatures in sf_by_feature.items():
        feature_fill = T2_FEATURE_PALETTE[feature_color_idx % len(T2_FEATURE_PALETTE)]
        feature_color_idx += 1
        
        for sf in feature_subfeatures:
            ws.cell(row=start_row, column=1).value = feature_name
            t2_style_cell(ws.cell(row=start_row, column=1), fill=feature_fill, alignment=T2_ALIGN_LEFT)
            
            ws.cell(row=start_row, column=2).value = sf.get("name", "")
            t2_style_cell(ws.cell(row=start_row, column=2), alignment=T2_ALIGN_LEFT)
            
            for col_idx, tool in enumerate(all_tools, start=3):
                cell_data = next(
                    (c for c in matrix_cells
                     if c["domain_subfeature_id"] == sf["id"] and c["domain_tool_id"] == tool["id"]),
                    None
                )
                
                support = cell_data.get("support_level", SUPPORT_NONE) if cell_data else SUPPORT_NONE
                
                cell = ws.cell(row=start_row, column=col_idx)
                cell.value = support
                t2_style_cell(
                    cell,
                    font=T2_FONT_SUPPORT,
                    fill=t2_get_support_fill(support),
                    alignment=T2_ALIGN_CENTER
                )
            
            start_row += 1


def _write_legend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Legend")
    
    headers = ["Symbol", "Meaning"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = label
        t2_style_cell(cell, font=T2_FONT_HEADER, fill=T2_FILL_TITLE, alignment=T2_ALIGN_CENTER)
    
    entries = [
        (SUPPORT_FULL, T2_FILL_GREEN, "Fully supported natively"),
        (SUPPORT_NONE, T2_FILL_RED, "Not supported"),
        (SUPPORT_PARTIAL, T2_FILL_AMBER, "Partial support (requires plugins/config)"),
    ]
    
    for row_idx, (symbol, fill, meaning) in enumerate(entries, start=2):
        ws.cell(row=row_idx, column=1).value = symbol
        t2_style_cell(ws.cell(row=row_idx, column=1), font=T2_FONT_SUPPORT, fill=fill, alignment=T2_ALIGN_CENTER)
        
        ws.cell(row=row_idx, column=2).value = meaning
        t2_style_cell(ws.cell(row=row_idx, column=2), font=T2_FONT_LEGEND, fill=fill, alignment=T2_ALIGN_LEFT)
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45


def _write_methodology_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Methodology")
    
    ws.cell(row=1, column=1).value = "Technique 2: Ranking Methodology"
    t2_style_cell(ws.cell(row=1, column=1), font=T2_FONT_TITLE, fill=T2_FILL_TITLE, alignment=T2_ALIGN_LEFT)
    
    methodology = [
        "",
        "COMPOSITE SCORE FORMULA:",
        "",
        f"Composite Score = {settings.t2_weight_subdomain_presence*100:.0f}% × Subdomain Presence + "
        f"{settings.t2_weight_feature_coverage*100:.0f}% × Feature Coverage + "
        f"{settings.t2_weight_market_presence*100:.0f}% × Market Presence + "
        f"{settings.t2_weight_rank_distribution*100:.0f}% × Rank Distribution",
        "",
        "FACTOR DEFINITIONS:",
        "",
        f"1. Subdomain Presence ({settings.t2_weight_subdomain_presence*100:.0f}%):",
        "   Percentage of subdomains within the domain that mention this tool.",
        "   Higher = more relevant across the domain.",
        "",
        f"2. Feature Coverage ({settings.t2_weight_feature_coverage*100:.0f}%):",
        "   Average support level across all features (✔=1.0, Partial=0.5, ✘=0.0).",
        "   Higher = better overall feature support.",
        "",
        f"3. Market Presence ({settings.t2_weight_market_presence*100:.0f}%):",
        "   LLM-assessed market leadership score (0-100).",
        "   Based on brand recognition, market share, analyst ratings.",
        "",
        f"4. Rank Distribution ({settings.t2_weight_rank_distribution*100:.0f}%):",
        "   Ratio of ✔ vs ✘ in subdomain matrices from Technique 1.",
        "   Higher = more fully-supported features.",
        "",
        "TOOL SELECTION:",
        f"- Top {settings.t2_max_enterprise_tools} enterprise tools (ranked 1-{settings.t2_max_enterprise_tools})",
        f"- Top {settings.t2_max_opensource_tools} open-source tools (ranked 1-{settings.t2_max_opensource_tools})",
    ]
    
    for row_idx, text in enumerate(methodology, start=2):
        ws.cell(row=row_idx, column=1).value = text
        t2_style_cell(ws.cell(row=row_idx, column=1), alignment=T2_ALIGN_LEFT, border=None)
    
    ws.column_dimensions["A"].width = 80


def _sync_export_domain_rankings(path: Path, rankings_data: list[dict]) -> None:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    _write_summary_sheet(wb, rankings_data)
    
    for domain_data in rankings_data:
        domain_name = domain_data.get("domain_name", "")
        tools_ent = domain_data.get("tools_enterprise", [])
        tools_oss = domain_data.get("tools_opensource", [])
        features = domain_data.get("features", [])
        subfeatures = domain_data.get("subfeatures", [])
        matrix_cells = domain_data.get("matrix_cells", [])
        
        _write_domain_sheet(
            wb,
            domain_name,
            tools_ent,
            tools_oss,
            features,
            subfeatures,
            matrix_cells,
        )
    
    _write_legend_sheet(wb)
    _write_methodology_sheet(wb)
    
    wb.save(str(path))
    logger.info(f"Technique 2 Excel exported to {path}")


async def export_domain_ranking_excel(domain_id: int, output_path: str | None = None) -> str:
    path = Path(output_path or settings.t2_excel_output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    
    ranking = await get_t2_domain_ranking(domain_id)
    if not ranking:
        raise ValueError(f"No ranking found for domain_id={domain_id}")
    
    tools_ent = await get_t2_domain_tools_by_type(domain_id, "enterprise")
    tools_oss = await get_t2_domain_tools_by_type(domain_id, "opensource")
    features = await get_t2_domain_features(domain_id)
    subfeatures = await get_t2_domain_subfeatures_by_domain(domain_id)
    matrix_cells = await get_t2_domain_matrix_cells(domain_id)
    
    domain_data = {
        "domain_name": ranking.get("domain_name", f"Domain {domain_id}"),
        "status": ranking.get("status", "pending"),
        "total_enterprise_tools": ranking.get("total_enterprise_tools", 0),
        "total_opensource_tools": ranking.get("total_opensource_tools", 0),
        "selected_enterprise_tools": ranking.get("selected_enterprise_tools", 0),
        "selected_opensource_tools": ranking.get("selected_opensource_tools", 0),
        "top_enterprise": {"name": tools_ent[0]["product_name"], "score": tools_ent[0]["composite_score"]} if tools_ent else {},
        "top_opensource": {"name": tools_oss[0]["product_name"], "score": tools_oss[0]["composite_score"]} if tools_oss else {},
        "tools_enterprise": tools_ent,
        "tools_opensource": tools_oss,
        "features": features,
        "subfeatures": subfeatures,
        "matrix_cells": matrix_cells,
    }
    
    async with _get_excel_lock():
        await asyncio.get_running_loop().run_in_executor(
            None,
            _sync_export_domain_rankings,
            path,
            [domain_data],
        )
    
    return str(path)


async def export_all_domain_rankings(output_path: str | None = None) -> str:
    path = Path(output_path or settings.t2_excel_output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    
    rankings = await get_all_t2_domain_rankings()
    
    if not rankings:
        logger.warning("No domain rankings found to export")
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        ws = wb.create_sheet("Summary")
        ws.cell(row=1, column=1).value = "No rankings available. Run Technique 2 pipeline first."
        wb.save(str(path))
        return str(path)
    
    rankings_data = []
    
    for ranking in rankings:
        domain_id = ranking.get("domain_id")
        
        tools_ent = await get_t2_domain_tools_by_type(domain_id, "enterprise")
        tools_oss = await get_t2_domain_tools_by_type(domain_id, "opensource")
        features = await get_t2_domain_features(domain_id)
        subfeatures = await get_t2_domain_subfeatures_by_domain(domain_id)
        matrix_cells = await get_t2_domain_matrix_cells(domain_id)
        
        rankings_data.append({
            "domain_name": ranking.get("domain_name", f"Domain {domain_id}"),
            "status": ranking.get("status", "pending"),
            "total_enterprise_tools": ranking.get("total_enterprise_tools", 0),
            "total_opensource_tools": ranking.get("total_opensource_tools", 0),
            "selected_enterprise_tools": ranking.get("selected_enterprise_tools", 0),
            "selected_opensource_tools": ranking.get("selected_opensource_tools", 0),
            "top_enterprise": {"name": tools_ent[0]["product_name"], "score": tools_ent[0]["composite_score"]} if tools_ent else {},
            "top_opensource": {"name": tools_oss[0]["product_name"], "score": tools_oss[0]["composite_score"]} if tools_oss else {},
            "tools_enterprise": tools_ent,
            "tools_opensource": tools_oss,
            "features": features,
            "subfeatures": subfeatures,
            "matrix_cells": matrix_cells,
        })
    
    async with _get_excel_lock():
        await asyncio.get_running_loop().run_in_executor(
            None,
            _sync_export_domain_rankings,
            path,
            rankings_data,
        )
    
    return str(path)


async def get_t2_domain_tools_by_type(domain_id: int, tool_type: str) -> list[dict]:
    from db.domain_store import get_t2_domain_tools
    
    tools = await get_t2_domain_tools(domain_id)
    return [t for t in tools if t.get("tool_type") == tool_type]
