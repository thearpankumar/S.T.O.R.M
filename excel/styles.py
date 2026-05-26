from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Support symbols ──────────────────────────────────────────────────────────
SUPPORT_FULL = "✔"
SUPPORT_NONE = "✘"
SUPPORT_PARTIAL = "Partial"


# ── Support cell fills (match sample exactly) ────────────────────────────────
FILL_GREEN   = PatternFill(start_color="D6F0E4", end_color="D6F0E4", fill_type="solid")
FILL_RED     = PatternFill(start_color="FAE0DC", end_color="FAE0DC", fill_type="solid")
FILL_AMBER   = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

# ── Header fills ─────────────────────────────────────────────────────────────
FILL_TITLE        = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
FILL_ENTERPRISE   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_OPENSOURCE   = PatternFill(start_color="375623", end_color="375623", fill_type="solid")

# ── Data column fills ────────────────────────────────────────────────────────
FILL_SUBDOMAIN  = PatternFill(start_color="EAF0F6", end_color="EAF0F6", fill_type="solid")
FILL_SUBFEATURE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# ── Rotating feature-group palette for column B ──────────────────────────────
FEATURE_PALETTE: list[PatternFill] = [
    PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid"),  # lavender
    PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),  # mint
    PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),  # warm yellow
    PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),  # sky blue
    PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),  # rose
    PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),  # lilac
    PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid"),  # cyan
]


# ── Fonts ────────────────────────────────────────────────────────────────────
FONT_TITLE      = Font(bold=True, color="FFFFFF", size=12)
FONT_ENT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_OSS_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_COL_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_SUBDOMAIN  = Font(bold=True, color="000000", size=10)
FONT_FEATURE    = Font(bold=True, color="000000", size=10)
FONT_SUBFEATURE = Font(bold=False, color="000000", size=10)
FONT_SUPPORT    = Font(bold=True,  color="000000", size=10)
FONT_LEGEND_HDR = Font(bold=True,  color="FFFFFF", size=10)
FONT_LEGEND_VAL = Font(bold=True,  color="000000", size=10)


# ── Alignments ───────────────────────────────────────────────────────────────
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── Borders ──────────────────────────────────────────────────────────────────
_THIN = Side(style="thin")
BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ── Helpers ──────────────────────────────────────────────────────────────────
def get_support_fill(value: str) -> PatternFill:
    return {
        SUPPORT_FULL:    FILL_GREEN,
        SUPPORT_NONE:    FILL_RED,
        SUPPORT_PARTIAL: FILL_AMBER,
    }.get(value, FILL_RED)


def style_cell(cell, *, font=None, fill=None, alignment=None, border=BORDER_THIN) -> None:
    """Apply non-None style attributes to a cell."""
    if font      is not None: cell.font      = font
    if fill      is not None: cell.fill      = fill
    if alignment is not None: cell.alignment = alignment
    if border    is not None: cell.border    = border


def set_column_widths(ws, num_tool_cols: int) -> None:
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 36
    for i in range(4, 4 + num_tool_cols):
        ws.column_dimensions[get_column_letter(i)].width = 12
