import asyncio
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from models.matrix import MatrixBatch, ToolSupportRow
from config.settings import settings
from excel.styles import (
    SUPPORT_FULL, SUPPORT_NONE, SUPPORT_PARTIAL,
    FILL_TITLE, FILL_ENTERPRISE, FILL_OPENSOURCE,
    FILL_SUBDOMAIN, FILL_SUBFEATURE,
    FILL_GREEN, FILL_RED, FILL_AMBER,
    FEATURE_PALETTE,
    FONT_TITLE, FONT_ENT_HEADER, FONT_OSS_HEADER, FONT_COL_HEADER,
    FONT_SUBDOMAIN, FONT_FEATURE, FONT_SUBFEATURE, FONT_SUPPORT,
    FONT_LEGEND_HDR, FONT_LEGEND_VAL,
    ALIGN_CENTER, ALIGN_LEFT,
    BORDER_THIN,
    get_support_fill, style_cell, set_column_widths,
)
import logging

logger = logging.getLogger(__name__)

_excel_lock: asyncio.Lock | None = None


def _get_excel_lock() -> asyncio.Lock:
    global _excel_lock
    if _excel_lock is None:
        _excel_lock = asyncio.Lock()
    return _excel_lock


# ── Low-level sheet writers ──────────────────────────────────────────────────

def _write_header_row_1(
    ws,
    title: str,
    tools_enterprise: list[dict[str, Any]],
    tools_opensource: list[dict[str, Any]],
) -> None:
    """Row 1: title span + Enterprise/Open-Source group headers."""
    num_enterprise = len(tools_enterprise)
    num_opensource = len(tools_opensource)

    # Title cell A1:C1
    ws.merge_cells("A1:C1")
    cell = ws["A1"]
    cell.value = title
    style_cell(cell, font=FONT_TITLE, fill=FILL_TITLE, alignment=ALIGN_CENTER)

    # Enterprise group header
    if num_enterprise > 0:
        ent_start = 4
        ent_end = ent_start + num_enterprise - 1
        ws.merge_cells(start_row=1, start_column=ent_start, end_row=1, end_column=ent_end)
        cell = ws.cell(row=1, column=ent_start)
        cell.value = "Enterprise Tools"
        style_cell(cell, font=FONT_ENT_HEADER, fill=FILL_ENTERPRISE, alignment=ALIGN_CENTER)
        # Style the merged-over cells too (borders)
        for col in range(ent_start + 1, ent_end + 1):
            style_cell(ws.cell(row=1, column=col), fill=FILL_ENTERPRISE, border=BORDER_THIN)

    # Open-Source group header
    if num_opensource > 0:
        oss_start = 4 + num_enterprise
        oss_end = oss_start + num_opensource - 1
        ws.merge_cells(start_row=1, start_column=oss_start, end_row=1, end_column=oss_end)
        cell = ws.cell(row=1, column=oss_start)
        cell.value = "Open-Source Tools"
        style_cell(cell, font=FONT_OSS_HEADER, fill=FILL_OPENSOURCE, alignment=ALIGN_CENTER)
        for col in range(oss_start + 1, oss_end + 1):
            style_cell(ws.cell(row=1, column=col), fill=FILL_OPENSOURCE, border=BORDER_THIN)

    # Style A1:C1 fill on merged-over B1/C1
    for col in (2, 3):
        style_cell(ws.cell(row=1, column=col), fill=FILL_TITLE, border=BORDER_THIN)

    ws.row_dimensions[1].height = 30


def _write_header_row_2(
    ws,
    tools_enterprise: list[dict[str, Any]],
    tools_opensource: list[dict[str, Any]],
) -> None:
    """Row 2: column headers (Subdomain / Feature / Sub-feature + tool names)."""
    for col_idx, label in enumerate(["Subdomain", "Feature", "Sub-feature"], start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = label
        style_cell(cell, font=FONT_COL_HEADER, fill=FILL_TITLE, alignment=ALIGN_CENTER)

    col = 4
    for tool in tools_enterprise:
        cell = ws.cell(row=2, column=col)
        cell.value = f"{tool.get('product_name', '')}\n{tool.get('vendor', '')}"
        style_cell(cell, font=FONT_ENT_HEADER, fill=FILL_ENTERPRISE, alignment=ALIGN_CENTER)
        col += 1

    for tool in tools_opensource:
        cell = ws.cell(row=2, column=col)
        cell.value = f"{tool.get('product_name', '')}\n{tool.get('vendor', '')}"
        style_cell(cell, font=FONT_OSS_HEADER, fill=FILL_OPENSOURCE, alignment=ALIGN_CENTER)
        col += 1

    ws.row_dimensions[2].height = 45


def _write_data_rows(
    ws,
    rows: list,
    tools_enterprise: list[dict[str, Any]],
    tools_opensource: list[dict[str, Any]],
) -> None:
    """Data rows: write + style all cells; merge A & B per feature group."""
    all_tools = (
        [t.get("product_name", "") for t in tools_enterprise]
        + [t.get("product_name", "") for t in tools_opensource]
    )
    total_cols = 3 + len(all_tools)

    # Group rows by feature (preserve insertion order)
    feature_groups: dict[str, list] = {}
    for row in rows:
        feature = row.feature if hasattr(row, "feature") else row.get("feature", "")
        feature_groups.setdefault(feature, []).append(row)

    row_idx = 3
    feature_color_idx = 0

    for feature, feature_rows in feature_groups.items():
        feature_fill = FEATURE_PALETTE[feature_color_idx % len(FEATURE_PALETTE)]
        feature_color_idx += 1
        feature_start_row = row_idx

        for row_data in feature_rows:
            subdomain   = row_data.subdomain   if hasattr(row_data, "subdomain")   else row_data.get("subdomain", "")
            feat        = row_data.feature     if hasattr(row_data, "feature")     else row_data.get("feature", "")
            sub_feat    = row_data.sub_feature if hasattr(row_data, "sub_feature") else row_data.get("sub_feature", "")
            tool_support = row_data.tool_support if hasattr(row_data, "tool_support") else row_data.get("tool_support", {})

            # Col A – subdomain
            cell_a = ws.cell(row=row_idx, column=1)
            cell_a.value = subdomain
            style_cell(cell_a, font=FONT_SUBDOMAIN, fill=FILL_SUBDOMAIN, alignment=ALIGN_CENTER)

            # Col B – feature
            cell_b = ws.cell(row=row_idx, column=2)
            cell_b.value = feat
            style_cell(cell_b, font=FONT_FEATURE, fill=feature_fill, alignment=ALIGN_LEFT)

            # Col C – sub-feature
            cell_c = ws.cell(row=row_idx, column=3)
            cell_c.value = sub_feat
            style_cell(cell_c, font=FONT_SUBFEATURE, fill=FILL_SUBFEATURE, alignment=ALIGN_LEFT)

            # Support columns
            col = 4
            for tool_name in all_tools:
                support = tool_support.get(tool_name, SUPPORT_NONE)
                cell = ws.cell(row=row_idx, column=col)
                cell.value = support
                style_cell(
                    cell,
                    font=FONT_SUPPORT,
                    fill=get_support_fill(support),
                    alignment=ALIGN_CENTER,
                )
                col += 1

            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

        # Merge A and B vertically across this feature group
        if len(feature_rows) > 1:
            ws.merge_cells(
                start_row=feature_start_row, start_column=1,
                end_row=row_idx - 1,         end_column=1,
            )
            ws.merge_cells(
                start_row=feature_start_row, start_column=2,
                end_row=row_idx - 1,         end_column=2,
            )
        # Even for single-row groups, re-center merged cell value
        ws.cell(row=feature_start_row, column=1).alignment = ALIGN_CENTER
        ws.cell(row=feature_start_row, column=2).alignment = ALIGN_LEFT

    set_column_widths(ws, len(all_tools))


def _ensure_legend_sheet(wb: Workbook) -> None:
    if "Legend" in wb.sheetnames:
        return

    ws = wb.create_sheet("Legend")

    # Header row
    for col_idx, label in enumerate(["Symbol", "Meaning"], start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = label
        style_cell(cell, font=FONT_LEGEND_HDR, fill=FILL_TITLE, alignment=ALIGN_CENTER)

    # Legend entries
    entries = [
        (SUPPORT_FULL,    FILL_GREEN, "Fully supported natively"),
        (SUPPORT_NONE,    FILL_RED,   "Not supported"),
        (SUPPORT_PARTIAL, FILL_AMBER, "Limited support / add-on or custom config required"),
    ]
    for row_idx, (symbol, fill, meaning) in enumerate(entries, start=2):
        sym_cell = ws.cell(row=row_idx, column=1)
        sym_cell.value = symbol
        style_cell(sym_cell, font=FONT_LEGEND_VAL, fill=fill, alignment=ALIGN_CENTER)

        mean_cell = ws.cell(row=row_idx, column=2)
        mean_cell.value = meaning
        style_cell(mean_cell, font=FONT_LEGEND_VAL, fill=fill, alignment=ALIGN_LEFT)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50


# ── Public per-batch writer ──────────────────────────────────────────────────

def _sync_write_excel(
    path: Path,
    batch: MatrixBatch,
    tools_enterprise: list[dict[str, Any]],
    tools_opensource: list[dict[str, Any]],
) -> None:
    if path.exists():
        wb = load_workbook(str(path))
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_name = f"{batch.subdomain} Matrix"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    _write_header_row_1(ws, f"{batch.subdomain} Matrix", tools_enterprise, tools_opensource)
    _write_header_row_2(ws, tools_enterprise, tools_opensource)
    _write_data_rows(ws, batch.rows, tools_enterprise, tools_opensource)

    _ensure_legend_sheet(wb)

    wb.save(str(path))
    logger.info(f"Saved Excel workbook to {path}")


async def write_matrix_batch(
    batch: MatrixBatch,
    tools_enterprise: list[dict[str, Any]],
    tools_opensource: list[dict[str, Any]],
    workbook_path: str | None = None,
) -> None:
    if not tools_enterprise and not tools_opensource:
        logger.warning("No tools to write, skipping Excel generation")
        return

    path = Path(workbook_path or settings.excel_output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    async with _get_excel_lock():
        await asyncio.get_running_loop().run_in_executor(
            None,
            _sync_write_excel,
            path,
            batch,
            tools_enterprise,
            tools_opensource,
        )


# ── DB-backed per-subdomain export ───────────────────────────────────────────

async def create_workbook_from_db(subdomain_id: int, subdomain_name: str) -> None:
    from db.store import get_tools, get_features, get_subfeatures, get_matrix_cells

    tools = await get_tools(subdomain_id)
    tools_enterprise = [t for t in tools if t["tool_type"] == "enterprise"]
    tools_opensource = [t for t in tools if t["tool_type"] == "opensource"]

    features = await get_features(subdomain_id)
    cells = await get_matrix_cells(subdomain_id)

    rows = []
    for feature in features:
        subfeatures = await get_subfeatures(feature["id"])

        for sf in subfeatures:
            tool_support = {}
            for cell in cells:
                if cell["subfeature_id"] == sf["id"]:
                    tool = next((t for t in tools if t["id"] == cell["tool_id"]), None)
                    if tool:
                        tool_support[tool["product_name"]] = cell["support_level"]

            if tool_support:
                rows.append(ToolSupportRow(
                    subdomain=subdomain_name,
                    feature=feature["name"],
                    sub_feature=sf["name"],
                    tool_support=tool_support,
                ))

    batch = MatrixBatch(
        subdomain=subdomain_name,
        tools_enterprise=[t["product_name"] for t in tools_enterprise],
        tools_opensource=[t["product_name"] for t in tools_opensource],
        rows=rows,
    )

    await write_matrix_batch(batch, tools_enterprise, tools_opensource)


def _write_placeholder_row(ws, subdomain_name: str, num_tool_cols: int) -> None:
    """Write a single greyed-out placeholder row when there are tools but no feature data yet."""
    from openpyxl.styles import PatternFill as _PF
    grey_fill = _PF(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    from openpyxl.styles import Font as _Font
    grey_font = _Font(italic=True, color="999999", size=10)

    placeholders = [subdomain_name, "—", "(feature data not yet collected)"]
    for col_idx, val in enumerate(placeholders, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = val
        cell.font = grey_font
        cell.fill = grey_fill
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER_THIN

    for col_idx in range(4, 4 + num_tool_cols):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = "—"
        cell.font = grey_font
        cell.fill = grey_fill
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    set_column_widths(ws, num_tool_cols)


# ── Full export (all subdomains → one workbook) ───────────────────────────────

def _sync_export_all(
    path: Path,
    subdomains_data: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary", 0)
    summary_headers = ["Domain", "Subdomain", "Status", "Tools", "Features"]
    for col_idx, label in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.value = label
        style_cell(cell, font=FONT_COL_HEADER, fill=FILL_TITLE, alignment=ALIGN_CENTER)

    row_idx = 2
    for sd_data in subdomains_data:
        ws_summary.cell(row=row_idx, column=1).value = sd_data.get("domain", "")
        ws_summary.cell(row=row_idx, column=2).value = sd_data.get("name", "")
        ws_summary.cell(row=row_idx, column=3).value = sd_data.get("status", "")
        ws_summary.cell(row=row_idx, column=4).value = sd_data.get("tools_count", 0)
        ws_summary.cell(row=row_idx, column=5).value = sd_data.get("features_count", 0)
        row_idx += 1

    for col in range(1, 6):
        ws_summary.column_dimensions[get_column_letter(col)].width = 30

    # ── Per-subdomain sheets (include any subdomain that has tools or features) ──
    for sd_data in subdomains_data:
        subdomain_name   = sd_data.get("name", "")
        tools_enterprise = sd_data.get("tools_enterprise", [])
        tools_opensource = sd_data.get("tools_opensource", [])
        rows             = sd_data.get("rows", [])
        status           = sd_data.get("status", "")

        # Skip subdomains that have absolutely no data at all
        has_tools = bool(tools_enterprise or tools_opensource)
        has_rows  = bool(rows)
        if not has_tools and not has_rows:
            continue

        sheet_name = subdomain_name[:31]
        ws = wb.create_sheet(sheet_name)

        # Add a "(Partial Data)" note in the title for non-done sheets
        title = f"{subdomain_name} Matrix"
        if status != "done":
            title += f" [{status.upper()}]"

        _write_header_row_1(ws, title, tools_enterprise, tools_opensource)
        _write_header_row_2(ws, tools_enterprise, tools_opensource)
        if rows:
            _write_data_rows(ws, rows, tools_enterprise, tools_opensource)
        elif has_tools:
            # No feature/matrix data yet — write a placeholder row
            _write_placeholder_row(ws, subdomain_name, len(tools_enterprise) + len(tools_opensource))

    _ensure_legend_sheet(wb)

    wb.save(str(path))
    logger.info(f"Exported all subdomains to {path}")


async def export_all_subdomains(output_path: str | None = None) -> str:
    from db.store import (
        get_domain_id, get_subdomains, get_tools,
        get_features, get_subfeatures, get_matrix_cells,
    )

    path = Path(output_path or settings.excel_output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    subdomains_data = []

    from config.domains import CYBERSECURITY_DOMAINS
    for domain in CYBERSECURITY_DOMAINS:
        domain_id = await get_domain_id(domain)
        if not domain_id:
            continue

        # Include ALL subdomains, not just 'done' ones
        subdomains = await get_subdomains(domain_id)
        for sd in subdomains:
            tools = await get_tools(sd["id"])
            tools_enterprise = [t for t in tools if t["tool_type"] == "enterprise"]
            tools_opensource = [t for t in tools if t["tool_type"] == "opensource"]

            features = await get_features(sd["id"])
            cells    = await get_matrix_cells(sd["id"])

            # Build rows from whatever data exists.
            # Sub-features with no matrix cells still get a row with all ✘.
            all_tool_names = (
                [t["product_name"] for t in tools_enterprise]
                + [t["product_name"] for t in tools_opensource]
            )
            rows = []
            for feature in features:
                subfeatures = await get_subfeatures(feature["id"])
                for sf in subfeatures:
                    # Start with ✘ for every tool, then fill in what the DB knows
                    tool_support: dict[str, str] = {
                        name: SUPPORT_NONE for name in all_tool_names
                    }
                    for cell in cells:
                        if cell["subfeature_id"] == sf["id"]:
                            tool = next(
                                (t for t in tools if t["id"] == cell["tool_id"]), None
                            )
                            if tool:
                                tool_support[tool["product_name"]] = cell["support_level"]

                    rows.append({
                        "subdomain":    sd["name"],
                        "feature":      feature["name"],
                        "sub_feature":  sf["name"],
                        "tool_support": tool_support,
                    })

            subdomains_data.append({
                "domain":         domain,
                "name":           sd["name"],
                "status":         sd["status"],
                "tools_count":    len(tools),
                "features_count": len(features),
                "tools_enterprise": [
                    {"product_name": t["product_name"], "vendor": t["vendor"]}
                    for t in tools_enterprise
                ],
                "tools_opensource": [
                    {"product_name": t["product_name"], "vendor": t["vendor"]}
                    for t in tools_opensource
                ],
                "rows": rows,
            })

    async with _get_excel_lock():
        await asyncio.get_running_loop().run_in_executor(
            None,
            _sync_export_all,
            path,
            subdomains_data,
        )

    return str(path)
