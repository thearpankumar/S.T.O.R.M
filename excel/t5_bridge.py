"""
excel/t5_bridge.py - Technique 5 Excel export.

Produces a 6-sheet workbook:
  Sheet 1 "Score Card" - Master list ranked by composite score
  Sheet 2 "Top Performers" - Top N tools highlight
  Sheet 3 "Dimension Analysis" - Chart breakdowns per dimension
  Sheet 4 "Grade Distribution" - Breakdown by letter grade
  Sheet 5 "Head-to-Head" - Side-by-side of top 10 tools
  Sheet 6 "Executive Dashboard" - Global KPIs
"""

import asyncio
import logging
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference, RadarChart
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import settings
from db.t5_store import get_t5_scores, get_t5_stats

logger = logging.getLogger(__name__)

# ── Styles ──
_DARK_BG = "0E1117"
_HEADER_BG = "1E2533"
_ROW_ALT1 = "131720"
_ROW_ALT2 = "1A1F2E"
_TEXT_MAIN = "F8FAFC"
_TEXT_DIM = "94A3B8"
_ACCENT = "3B82F6"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color: str = _TEXT_MAIN, bold: bool = False, size: int = 10) -> Font:
    return Font(color=hex_color, bold=bold, size=size)


def _align(horizontal: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def _thin_border() -> Border:
    side = Side(style="thin", color="2D3748")
    return Border(left=side, right=side, top=side, bottom=side)


def _style(cell, fill_hex: str | None = None, font: Font | None = None,
           alignment: Alignment | None = None, border: bool = True) -> None:
    if fill_hex:
        cell.fill = _fill(fill_hex)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = _thin_border()


def _get_grade_color(grade: str) -> str:
    return {
        "A+": "15803D", "A": "16A34A",
        "B+": "CA8A04", "B": "EAB308",
        "C":  "EA580C", "D": "B91C1C"
    }.get(grade, "94A3B8")


# ── Sheet 1: Score Card ──
def _write_scorecard_sheet(wb: Workbook, scores: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Score Card")
    
    ws.merge_cells("A1:N1")
    title = ws["A1"]
    title.value = "Technique 5 - Master Tool Score Card"
    _style(title, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=14),
           alignment=_align("center"))
    ws.row_dimensions[1].height = 32
    
    headers = ["Domain", "Domain Rank", "Vendor", "Product Name", "Category", "Grade", "Composite",
               "D1 (Feature)", "D2 (Domain)", "D3 (NIST)", "D4 (Market)", "D5 (Rank)",
               "Quadrant", "Strategic Insight"]
    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        _style(cell, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True),
               alignment=_align("center"))
    ws.row_dimensions[2].height = 20
    
    for i, t in enumerate(scores, 1):
        row_idx = i + 2
        fill_hex = _ROW_ALT1 if i % 2 == 1 else _ROW_ALT2
        
        grade = t.get("grade", "-")
        grade_color = _get_grade_color(grade)
        
        values = [
            t.get("primary_domain", ""),
            t.get("domain_rank", i),
            t.get("vendor", ""),
            t.get("product_name", ""),
            t.get("tool_category", ""),
            grade,
            f"{t.get('composite_score', 0):.1f}",
            f"{t.get('d1_feature_coverage', 0):.1f}",
            f"{t.get('d2_domain_breadth', 0):.1f}",
            f"{t.get('d3_nist_alignment', 0):.1f}",
            f"{t.get('d4_market_maturity', 0):.1f}",
            f"{t.get('d5_ranking_signal', 0):.1f}",
            t.get("quadrant_position", ""),
            t.get("strategic_insight", "")
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            align = _align("center") if col_idx in (2, 5, 6, 7, 8, 9, 10, 11, 12, 13) else _align("left")
            
            if col_idx == 6: # Grade
                _style(cell, fill_hex=grade_color, font=_font(_TEXT_MAIN, bold=True), alignment=align)
            elif col_idx == 7: # Composite
                _style(cell, fill_hex=fill_hex, font=_font(_TEXT_MAIN, bold=True), alignment=align)
            else:
                _style(cell, fill_hex=fill_hex, font=_font(_TEXT_MAIN), alignment=align)
                
        ws.row_dimensions[row_idx].height = 18
        
    widths = [20, 12, 22, 32, 14, 8, 12, 12, 12, 12, 12, 12, 16, 60]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:N{len(scores) + 2}"


def _write_domain_scorecards_sheet(wb: Workbook, scores: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Domain Scorecards")
    
    # Group by domain
    from collections import defaultdict
    domains = defaultdict(list)
    for t in scores:
        domains[t.get("primary_domain", "Unknown")].append(t)
    
    row_idx = 2
    for domain, d_scores in domains.items():
        # Write domain header
        ws.merge_cells(f"B{row_idx}:I{row_idx}")
        cell = ws[f"B{row_idx}"]
        cell.value = f"Domain: {domain}"
        _style(cell, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=14), alignment=_align("left"))
        row_idx += 1
        
        # Write column headers
        headers = ["Rank", "Vendor", "Product", "Category", "Grade", "Composite", "Quadrant", "Strategic Insight"]
        for col_idx, label in enumerate(headers, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=label)
            _style(cell, fill_hex=_ROW_ALT2, font=_font(_TEXT_MAIN, bold=True), alignment=_align("center"))
        row_idx += 1
        
        for t in d_scores[:10]: # Top 10 per domain
            fill_hex = _ROW_ALT1 if row_idx % 2 == 1 else _ROW_ALT2
            values = [
                t.get("domain_rank", "-"),
                t.get("vendor", ""),
                t.get("product_name", ""),
                t.get("tool_category", ""),
                t.get("grade", "-"),
                f"{t.get('composite_score', 0):.1f}",
                t.get("quadrant_position", ""),
                t.get("strategic_insight", "")
            ]
            for col_idx, value in enumerate(values, 2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                align = _align("center") if col_idx in (2, 5, 6, 7, 8) else _align("left")
                _style(cell, fill_hex=fill_hex, font=_font(_TEXT_MAIN), alignment=align)
                if col_idx == 6: # Grade
                    _style(cell, fill_hex=_get_grade_color(t.get("grade", "-")), font=_font(_TEXT_MAIN, bold=True), alignment=align)
            row_idx += 1
        row_idx += 2
        
    widths = [4, 6, 20, 28, 14, 8, 12, 16, 60]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


# ── Sheet 3 & 4 (Combined helper): Grade & Dimensions ──
def _write_grade_dist_sheet(wb: Workbook, stats: dict[str, Any]) -> None:
    ws = wb.create_sheet("Grade Distribution")
    ws.merge_cells("A1:C1")
    ws["A1"] = "Grade Distribution"
    _style(ws["A1"], fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=13), alignment=_align("center"))
    ws.row_dimensions[1].height = 28
    
    headers = ["Grade", "Count", "Percentage"]
    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        _style(cell, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True), alignment=_align("center"))
        
    grades = stats.get("grade_distribution", {})
    total = sum(grades.values())
    
    if total == 0:
        ws["A3"] = "No data"
        return
    
    row_idx = 3
    for grade, count in grades.items():
        if count == 0 and grade not in ("A+", "A", "B", "C"): continue
        fill_hex = _ROW_ALT1 if row_idx % 2 == 1 else _ROW_ALT2
        
        values = [grade, count, f"{(count/total)*100:.1f}%"]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _style(cell, fill_hex=fill_hex, font=_font(_TEXT_MAIN), alignment=_align("center"))
        row_idx += 1
        
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12


def _write_executive_dashboard(wb: Workbook, stats: dict[str, Any], scores: list[dict]) -> None:
    ws = wb.create_sheet("Executive Dashboard")
    ws.sheet_properties.tabColor = "3B82F6"
    ws.sheet_view.showGridLines = False
    
    for r in range(1, 35):
        for c in range(1, 18):
            _style(ws.cell(row=r, column=c), fill_hex=_DARK_BG, border=False)
            
    ws.merge_cells("B2:Q3")
    title = ws["B2"]
    title.value = "Technique 5: Strategic Tool Score Card"
    _style(title, fill_hex=_DARK_BG, font=_font(_TEXT_MAIN, bold=True, size=20), alignment=_align("left"))
    
    top_tool = scores[0] if scores else {}
    
    ws.merge_cells("B5:H10")
    summary = ws["B5"]
    summary.value = (
        f"Total Tools Scored: {stats.get('total', 0)}\n"
        f"Average Composite Score: {stats.get('avg_composite') or 0.0:.1f}/100\n"
        f"A-Grade Tools (A+/A): {stats.get('grade_distribution', {}).get('A+', 0) + stats.get('grade_distribution', {}).get('A', 0)}\n\n"
        f"Top Ranked Tool: {top_tool.get('vendor', '')} {top_tool.get('product_name', '')} ({top_tool.get('grade', '')})"
    )
    _style(summary, fill_hex=_ROW_ALT1, font=_font(_TEXT_DIM, size=11), alignment=_align("left", wrap=True), border=True)
    
    # Radar Chart Data for Global Averages
    data_ws = wb.create_sheet("_T5_DashboardData")
    data_ws.sheet_state = "hidden"
    
    dim_avgs = stats.get("dimension_averages", {})
    dims = ["D1: Feature", "D2: Domain", "D3: NIST", "D4: Market", "D5: Rank"]
    vals = [dim_avgs.get("d1", 0), dim_avgs.get("d2", 0), dim_avgs.get("d3", 0), dim_avgs.get("d4", 0), dim_avgs.get("d5", 0)]
    
    data_ws["A1"] = "Dimension"
    data_ws["B1"] = "Global Avg"
    for r, (d, v) in enumerate(zip(dims, vals), 2):
        data_ws[f"A{r}"] = d
        data_ws[f"B{r}"] = v
        
    radar = RadarChart()
    radar.type = "standard"
    labels = Reference(data_ws, min_col=1, min_row=2, max_row=6)
    data = Reference(data_ws, min_col=2, min_row=1, max_row=6)
    radar.add_data(data, titles_from_data=True)
    radar.set_categories(labels)
    radar.title = "Global Dimension Averages"
    radar.width = 16
    radar.height = 12
    ws.add_chart(radar, "K5")


def _sync_write_t5_excel(path: Path, scores: list[dict], stats: dict) -> None:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    _write_executive_dashboard(wb, stats, scores)
    _write_scorecard_sheet(wb, scores)
    _write_domain_scorecards_sheet(wb, scores)
    _write_grade_dist_sheet(wb, stats)
    
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    logger.info(f"T5 Excel workbook saved → {path}")


async def export_t5_workbook() -> str:
    scores = await get_t5_scores()
    if not scores:
        raise ValueError("No T5 scores found. Run T5 Score Card pipeline first.")
        
    stats = await get_t5_stats()
    output_path = Path(settings.t5_excel_output_path)
    
    # Run synchronously. Openpyxl with a few hundred tools writes in <0.1s.
    # We avoid run_in_executor to ensure it is not blocked by thread pool starvation.
    _sync_write_t5_excel(
        output_path,
        scores,
        stats
    )
    
    return str(output_path)
