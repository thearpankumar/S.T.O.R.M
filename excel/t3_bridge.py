"""
excel/t3_bridge.py — Technique 3 Excel export.

Produces a 4-sheet workbook:
  Sheet 1 "Classification"  — flat table, one row per unique tool
  Sheet 2 "NIST Coverage"   — tool × NIST function boolean matrix
  Sheet 3 "Domain Coverage" — tool × domain boolean matrix
  Sheet 4 "Summary"         — aggregate statistics
"""

import asyncio
import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import RadarChart, PieChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule

from config.settings import settings

logger = logging.getLogger(__name__)


# ── Colour palette ─────────────────────────────────────────────────────────────

_DARK_BG   = "0E1117"
_HEADER_BG = "1E2533"
_ROW_ALT1  = "131720"
_ROW_ALT2  = "1A1F2E"
_TEXT_MAIN = "F8FAFC"
_TEXT_DIM  = "94A3B8"
_ACCENT    = "3B82F6"

# NIST function colours (background fills)
_NIST_COLORS: dict[str, str] = {
    "ID": "1D4ED8",  # Blue   — Identify
    "PR": "15803D",  # Green  — Protect
    "DE": "C2410C",  # Orange — Detect
    "RS": "B91C1C",  # Red    — Respond
    "RC": "7E22CE",  # Purple — Recover
    "GV": "0E7490",  # Teal   — Govern
}

_NIST_LABELS = {
    "ID": "Identify",
    "PR": "Protect",
    "DE": "Detect",
    "RS": "Respond",
    "RC": "Recover",
    "GV": "Govern",
}

_CHECK = "✔"
_DASH  = "–"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(
    hex_color: str = _TEXT_MAIN,
    bold: bool = False,
    size: int = 10,
    italic: bool = False,
) -> Font:
    return Font(color=hex_color, bold=bold, size=size, italic=italic)


def _align(horizontal: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def _thin_border() -> Border:
    side = Side(style="thin", color="2D3748")
    return Border(left=side, right=side, top=side, bottom=side)


def _style(
    cell: Any,
    fill_hex: str | None = None,
    font: Font | None = None,
    alignment: Alignment | None = None,
    border: bool = True,
) -> None:
    if fill_hex:
        cell.fill = _fill(fill_hex)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = _thin_border()


# ── Sheet 1: Flat classification table ────────────────────────────────────────

def _write_classification_sheet(
    wb: Workbook,
    tools: list[dict[str, Any]],
    tool_memberships: dict[int, list[dict]],
) -> None:
    ws = wb.create_sheet("Classification Matrix")

    # ── Title row ──
    title_cols = 10
    ws.merge_cells(f"A1:{get_column_letter(title_cols)}1")
    tc = ws["A1"]
    tc.value = "Technique 3 — Cross-Domain Tool Classification"
    _style(tc, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=13), alignment=_align("center"))
    ws.row_dimensions[1].height = 28

    # ── Column headers ──
    headers = [
        "#", "Vendor", "Product Name", "Type",
        "NIST Primary", "NIST Functions", "# Domains", "Domain Names",
        "# Subdomains", "Subdomain Names",
    ]
    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        _style(cell, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=10), alignment=_align("center"))
    ws.row_dimensions[2].height = 20

    # ── Data rows ──
    for i, tool in enumerate(tools, 1):
        row_idx = i + 2
        fill_hex = _ROW_ALT1 if i % 2 == 1 else _ROW_ALT2

        memberships = tool_memberships.get(tool["id"], [])
        domain_names = sorted({m["domain_name"] for m in memberships})
        subdomain_names = sorted({m["subdomain_name"] for m in memberships})

        nist_raw = tool.get("nist_functions") or "[]"
        try:
            nist_list: list[str] = json.loads(nist_raw) if isinstance(nist_raw, str) else nist_raw
        except (json.JSONDecodeError, TypeError):
            nist_list = []
        nist_primary = tool.get("nist_primary_function") or ""
        nist_str = ", ".join(nist_list)

        values = [
            i,
            tool.get("vendor", ""),
            tool.get("product_name", ""),
            (tool.get("tool_type") or "").capitalize(),
            nist_primary,
            nist_str,
            tool.get("domain_count") or len(domain_names),
            ", ".join(domain_names),
            tool.get("subdomain_count") or len(subdomain_names),
            ", ".join(subdomain_names),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            font = _font(_TEXT_MAIN)
            align = _align("center") if col_idx in (1, 4, 5, 7, 9) else _align("left", wrap=(col_idx in (8, 10)))

            # NIST primary cell: colour by function
            if col_idx == 5 and nist_primary in _NIST_COLORS:
                _style(cell, fill_hex=_NIST_COLORS[nist_primary], font=_font(_TEXT_MAIN, bold=True), alignment=align)
            # Domain count: gradient green tint
            elif col_idx == 7:
                domain_cnt = tool.get("domain_count", 0)
                intensity = min(int(domain_cnt / 5 * 255), 200)
                green_hex = f"00{intensity:02X}40"
                _style(cell, fill_hex=fill_hex, font=_font(_TEXT_MAIN if domain_cnt <= 2 else "DCFCE7"), alignment=align)
            else:
                _style(cell, fill_hex=fill_hex, font=font, alignment=align)

        ws.row_dimensions[row_idx].height = 18

    # ── Column widths ──
    widths = [5, 22, 32, 12, 14, 26, 12, 55, 14, 75]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze panes & auto-filter ──
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(title_cols)}{len(tools) + 2}"


# ── Sheet 2: NIST Coverage Matrix ─────────────────────────────────────────────

def _write_nist_sheet(wb: Workbook, tools: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("NIST Coverage")
    nist_funcs = ["ID", "PR", "DE", "RS", "RC", "GV"]

    # Title
    ws.merge_cells(f"A1:{get_column_letter(2 + len(nist_funcs))}1")
    tc = ws["A1"]
    tc.value = "NIST CSF 2.0 Coverage Matrix — Tool × Function"
    _style(tc, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=12), alignment=_align("center"))
    ws.row_dimensions[1].height = 26

    # Header row
    ws.cell(row=2, column=1, value="Vendor")
    ws.cell(row=2, column=2, value="Product Name")
    _style(ws.cell(row=2, column=1), fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True), alignment=_align("center"))
    _style(ws.cell(row=2, column=2), fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True), alignment=_align("left"))

    for fi, func in enumerate(nist_funcs, 3):
        cell = ws.cell(row=2, column=fi, value=f"{func}\n{_NIST_LABELS[func]}")
        _style(cell, fill_hex=_NIST_COLORS[func], font=_font(_TEXT_MAIN, bold=True, size=9), alignment=_align("center", wrap=True))
        ws.column_dimensions[get_column_letter(fi)].width = 11
    ws.row_dimensions[2].height = 30

    # Data rows
    for i, tool in enumerate(tools, 1):
        row_idx = i + 2
        fill_hex = _ROW_ALT1 if i % 2 == 1 else _ROW_ALT2

        nist_raw = tool.get("nist_functions") or "[]"
        try:
            nist_set = set(json.loads(nist_raw) if isinstance(nist_raw, str) else nist_raw)
        except (json.JSONDecodeError, TypeError):
            nist_set = set()

        c1 = ws.cell(row=row_idx, column=1, value=tool.get("vendor", ""))
        _style(c1, fill_hex=fill_hex, font=_font(_TEXT_DIM, size=9), alignment=_align("left"))
        c2 = ws.cell(row=row_idx, column=2, value=tool.get("product_name", ""))
        _style(c2, fill_hex=fill_hex, font=_font(_TEXT_MAIN), alignment=_align("left"))

        for fi, func in enumerate(nist_funcs, 3):
            has_func = func in nist_set
            bg_color = "15803D" if has_func else fill_hex
            cell = ws.cell(row=row_idx, column=fi, value=_CHECK if has_func else _DASH)
            _style(cell, fill_hex=bg_color, font=_font(_TEXT_MAIN, bold=has_func), alignment=_align("center"))

        ws.row_dimensions[row_idx].height = 16

    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32


# ── Sheet 3: Domain Coverage Matrix ───────────────────────────────────────────

def _write_domain_sheet(
    wb: Workbook,
    tools: list[dict[str, Any]],
    tool_memberships: dict[int, list[dict]],
    all_domains: list[str],
) -> None:
    ws = wb.create_sheet("Domain Coverage")
    n_domains = len(all_domains)
    total_cols = 2 + n_domains

    # Title
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    tc = ws["A1"]
    tc.value = "Domain Coverage Matrix — Tool × Domain"
    _style(tc, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=12), alignment=_align("center"))
    ws.row_dimensions[1].height = 26

    # Headers
    ws.cell(row=2, column=1, value="Vendor")
    ws.cell(row=2, column=2, value="Product Name")
    _style(ws.cell(row=2, column=1), fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True), alignment=_align("center"))
    _style(ws.cell(row=2, column=2), fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True), alignment=_align("left"))

    # Short domain header labels (first 12 chars to keep columns narrow)
    for di, domain in enumerate(all_domains, 3):
        cell = ws.cell(row=2, column=di, value=domain[:15])
        _style(cell, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=8), alignment=_align("center", wrap=True))
        ws.column_dimensions[get_column_letter(di)].width = 12
    ws.row_dimensions[2].height = 38

    # Data rows
    for i, tool in enumerate(tools, 1):
        row_idx = i + 2
        fill_hex = _ROW_ALT1 if i % 2 == 1 else _ROW_ALT2
        tool_domains = {m["domain_name"] for m in tool_memberships.get(tool["id"], [])}

        c1 = ws.cell(row=row_idx, column=1, value=tool.get("vendor", ""))
        _style(c1, fill_hex=fill_hex, font=_font(_TEXT_DIM, size=9), alignment=_align("left"))
        c2 = ws.cell(row=row_idx, column=2, value=tool.get("product_name", ""))
        _style(c2, fill_hex=fill_hex, font=_font(_TEXT_MAIN), alignment=_align("left"))

        for di, domain in enumerate(all_domains, 3):
            has_domain = domain in tool_domains
            bg_color = "3B82F6" if has_domain else fill_hex
            cell = ws.cell(row=row_idx, column=di, value=_CHECK if has_domain else _DASH)
            _style(cell, fill_hex=bg_color, font=_font(_TEXT_MAIN, bold=has_domain), alignment=_align("center"))

        ws.row_dimensions[row_idx].height = 16

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32


# ── Sheet 4: Summary Statistics ───────────────────────────────────────────────

def _write_summary_sheet(
    wb: Workbook,
    stats: dict[str, Any],
    tools: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Summary")

    # Title
    ws.merge_cells("A1:C1")
    tc = ws["A1"]
    tc.value = "Technique 3 — Classification Summary"
    _style(tc, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True, size=13), alignment=_align("center"))
    ws.row_dimensions[1].height = 28

    # Metrics header
    for col, label in enumerate(["Metric", "Value", "Notes"], 1):
        cell = ws.cell(row=2, column=col, value=label)
        _style(cell, fill_hex=_HEADER_BG, font=_font(_TEXT_MAIN, bold=True), alignment=_align("center"))
    ws.row_dimensions[2].height = 18

    metrics: list[tuple[str, Any, str]] = [
        ("Total unique tools", stats.get("total", 0), "Across all completed T1 subdomains"),
        ("Enterprise tools", stats.get("enterprise", 0), ""),
        ("Open-source tools", stats.get("opensource", 0), ""),
        ("Tools in 1 domain only", stats.get("total", 0) - stats.get("multi_domain", 0), ""),
        ("Tools in 2+ domains", stats.get("multi_domain", 0), "Cross-domain tools"),
    ]

    top = stats.get("top_tool")
    if top:
        metrics.append((
            "Most cross-domain tool",
            f"{top.get('vendor','')} {top.get('product_name','')}",
            f"{top.get('domain_count', 0)} domains",
        ))

    metrics.append(("", "", ""))
    metrics.append(("── NIST Function Breakdown ──", "", ""))

    nist_counts = stats.get("nist_counts", {})
    for func in ("ID", "PR", "DE", "RS", "RC", "GV"):
        metrics.append((
            f"NIST {func} — {_NIST_LABELS.get(func, '')}",
            nist_counts.get(func, 0),
            f"Tools classified with {func} function",
        ))

    for i, (metric, value, note) in enumerate(metrics, 1):
        row_idx = i + 2
        fill_hex = _ROW_ALT1 if i % 2 == 1 else _ROW_ALT2

        is_section = str(metric).startswith("──")
        font = _font(_TEXT_MAIN, bold=True) if is_section else _font(_TEXT_MAIN)
        fill = _HEADER_BG if is_section else fill_hex

        c1 = ws.cell(row=row_idx, column=1, value=metric)
        _style(c1, fill_hex=fill, font=font, alignment=_align("left"))
        c2 = ws.cell(row=row_idx, column=2, value=value if not is_section else "")
        _style(c2, fill_hex=fill, font=_font(_TEXT_MAIN, bold=not is_section), alignment=_align("center"))
        c3 = ws.cell(row=row_idx, column=3, value=note if not is_section else "")
        _style(c3, fill_hex=fill, font=_font(_TEXT_DIM, italic=True, size=9), alignment=_align("left"))
        ws.row_dimensions[row_idx].height = 18

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 45


from openpyxl.chart import RadarChart, PieChart, BarChart, Reference, Series
from collections import Counter

def _write_executive_dashboard_sheet(wb: Workbook, exec_summary: str, stats: dict[str, Any], tools: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Executive Dashboard")
    ws.sheet_properties.tabColor = "3B82F6"
    
    # Hide gridlines
    ws.sheet_view.showGridLines = False
    
    # Background
    for r in range(1, 40):
        for c in range(1, 20):
            _style(ws.cell(row=r, column=c), fill_hex=_DARK_BG, border=False)

    # Title
    ws.merge_cells("B2:P3")
    title = ws["B2"]
    title.value = "Strategic Tool Landscape Analysis"
    _style(title, fill_hex=_DARK_BG, font=_font(_TEXT_MAIN, bold=True, size=20), alignment=_align("left"))
    
    # Summary Box
    ws.merge_cells("B5:P12")
    summary = ws["B5"]
    summary.value = exec_summary or "No agentic summary generated for this run."
    _style(summary, fill_hex=_ROW_ALT1, font=_font(_TEXT_DIM, size=11, italic=True), alignment=_align("left", wrap=True), border=True)

    # Calculate fallbacks in case `stats` is empty
    nist_counts = stats.get("nist_counts", {})
    if not any(nist_counts.values()):
        nist_counts = {"ID": 0, "PR": 0, "DE": 0, "RS": 0, "RC": 0, "GV": 0}
        for t in tools:
            nf = t.get("nist_functions")
            if nf:
                try:
                    for f in (json.loads(nf) if isinstance(nf, str) else nf):
                        if f in nist_counts: nist_counts[f] += 1
                except: pass

    ent_count = stats.get("enterprise", 0)
    os_count = stats.get("opensource", 0)
    if ent_count == 0 and os_count == 0:
        for t in tools:
            ttype = str(t.get("tool_type", "")).lower()
            if "commercial" in ttype or "enterprise" in ttype: ent_count += 1
            elif "open" in ttype: os_count += 1

    # Create hidden ChartData sheet to avoid blank charts
    data_ws = wb.create_sheet("_ChartData")
    data_ws.sheet_state = "hidden"

    # Radar Data
    data_ws['A1'] = 'Function'
    data_ws['B1'] = 'Count'
    row_idx = 2
    for fn in ("ID", "PR", "DE", "RS", "RC", "GV"):
        data_ws[f'A{row_idx}'] = fn
        data_ws[f'B{row_idx}'] = nist_counts.get(fn, 0)
        row_idx += 1
        
    # Pie Data
    data_ws['A10'] = 'Type'
    data_ws['B10'] = 'Count'
    data_ws['A11'] = 'Commercial'
    data_ws['B11'] = ent_count
    data_ws['A12'] = 'Open Source'
    data_ws['B12'] = os_count

    # Vendor Data
    vendor_counts = Counter([t.get("vendor", "Unknown") for t in tools])
    top_vendors = vendor_counts.most_common(10)
    data_ws['A15'] = 'Vendor'
    data_ws['B15'] = 'Count'
    row_idx = 16
    for v, c in top_vendors:
        data_ws[f'A{row_idx}'] = str(v)
        data_ws[f'B{row_idx}'] = int(c)
        row_idx += 1
        
    # Platform Data
    sub_counts = Counter([t.get("subdomain_count", 0) for t in tools])
    bins = {"1-2": 0, "3-5": 0, "6-10": 0, "11+": 0}
    for k, v in sub_counts.items():
        if k is None: k = 0
        if k <= 2: bins["1-2"] += v
        elif k <= 5: bins["3-5"] += v
        elif k <= 10: bins["6-10"] += v
        else: bins["11+"] += v
        
    data_ws['A30'] = 'Subdomains'
    data_ws['B30'] = 'Tools'
    row_idx = 31
    for k, v in bins.items():
        data_ws[f'A{row_idx}'] = k
        data_ws[f'B{row_idx}'] = int(v)
        row_idx += 1

    # Radar Chart
    radar = RadarChart()
    radar.plotVisOnly = False
    radar.type = "filled"
    radar.title = "NIST CSF 2.0 Coverage"
    radar.style = 26
    radar.width = 16
    radar.height = 10
    labels = Reference(data_ws, min_col=1, min_row=2, max_row=7)
    data = Reference(data_ws, min_col=2, min_row=1, max_row=7)
    radar.add_data(data, titles_from_data=True)
    radar.set_categories(labels)
    ws.add_chart(radar, "B14")

    # Pie Chart
    pie = PieChart()
    pie.plotVisOnly = False
    pie.title = "Tool Ecosystem Breakdown"
    pie.style = 10
    pie.width = 16
    pie.height = 10
    labels_pie = Reference(data_ws, min_col=1, min_row=11, max_row=12)
    data_pie = Reference(data_ws, min_col=2, min_row=10, max_row=12)
    pie.add_data(data_pie, titles_from_data=True)
    pie.set_categories(labels_pie)
    ws.add_chart(pie, "J14")

    # Bar Chart (Vendors)
    bar = BarChart()
    bar.type = "col"  # Force vertical columns to avoid overlapping axes!
    bar.plotVisOnly = False
    bar.title = "Top 10 Vendors by Footprint"
    bar.style = 10
    bar.width = 16
    bar.height = 10
    labels_bar = Reference(data_ws, min_col=1, min_row=16, max_row=25)
    data_bar = Reference(data_ws, min_col=2, min_row=15, max_row=25)
    bar.add_data(data_bar, titles_from_data=True)
    bar.set_categories(labels_bar)
    # Ensure X-axis labels are readable
    bar.x_axis.tickLblPos = "low"
    ws.add_chart(bar, "B35")
    
    # Bar Chart (Platform vs Point Solution)
    hist = BarChart()
    hist.type = "col"  # Force vertical columns!
    hist.plotVisOnly = False
    hist.title = "Platform vs Point Solution"
    hist.style = 13
    hist.width = 16
    hist.height = 10
    labels_hist = Reference(data_ws, min_col=1, min_row=31, max_row=34)
    data_hist = Reference(data_ws, min_col=2, min_row=30, max_row=34)
    hist.add_data(data_hist, titles_from_data=True)
    hist.set_categories(labels_hist)
    hist.x_axis.tickLblPos = "low"
    ws.add_chart(hist, "J35")


# ── Main sync writer ───────────────────────────────────────────────────────────

def _sync_write_t3_excel(
    path: Path,
    tools: list[dict[str, Any]],
    tool_memberships: dict[int, list[dict]],
    all_domains: list[str],
    stats: dict[str, Any],
    exec_summary: str,
) -> None:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Order sheets
    _write_executive_dashboard_sheet(wb, exec_summary, stats, tools)
    _write_classification_sheet(wb, tools, tool_memberships)
    _write_nist_sheet(wb, tools)
    _write_domain_sheet(wb, tools, tool_memberships, all_domains)
    _write_summary_sheet(wb, stats, tools)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    logger.info(f"T3 Excel workbook saved → {path} ({len(tools)} tools, 4 sheets)")


# ── Async public API ───────────────────────────────────────────────────────────

async def export_t3_workbook() -> str:
    """
    Export the full T3 classification workbook.
    Fetches data from the DB, writes 4-sheet Excel, returns the output path.
    """
    from db.t3_store import (
        get_t3_tools_with_coverage,
        get_all_t3_tool_memberships,
        get_t3_stats,
        get_t3_run_status,
    )
    from config.domains import CYBERSECURITY_DOMAINS

    tools = await get_t3_tools_with_coverage()

    if not tools:
        raise ValueError("No T3 classified tools found. Run the T3 classification pipeline first.")

    # Single query for all memberships — avoids N+1 at 2000+ tools
    tool_memberships = await get_all_t3_tool_memberships()

    all_domains = list(CYBERSECURITY_DOMAINS)
    stats = await get_t3_stats()
    
    t3_run = await get_t3_run_status()
    exec_summary = t3_run.get("executive_summary", "") if t3_run else ""

    output_path = Path(settings.t3_excel_output_path)

    # Use run_in_executor directly (no module-level lock needed — openpyxl write
    # is CPU-bound and single-threaded; the executor handles serialization).
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(
        None,
        _sync_write_t3_excel,
        output_path,
        tools,
        tool_memberships,
        all_domains,
        stats,
        exec_summary,
    )

    return str(output_path)
